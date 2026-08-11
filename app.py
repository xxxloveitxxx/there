import os
import time
import base64
import requests
import io
import json
from flask import abort, Flask, render_template, request, redirect, jsonify, make_response, url_for
from datetime import date, datetime, timezone, timedelta
from email.mime.text import MIMEText
from supabase import create_client, Client
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import Flow
from google.oauth2 import id_token
import google.auth.transport.requests as grequests
from flask_cors import CORS  
from cryptography.fernet import Fernet
from transaction_autopilot import bp as autopilot_bp
from public import public_bp
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import re
import dns.resolver
import csv
from io import TextIOWrapper
from openpyxl import load_workbook
from collections import defaultdict
from functools import wraps

# You need these imports for SMTP
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# ── single Flask app & blueprint registration ──
app = Flask(__name__, template_folder="templates")
CORS(app, resources={r"/connect-smtp": {"origins": "https://replyzeai.com"}})
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret")


# Fixed rate limit decorator
# Add these imports at the top
import supabase
from datetime import datetime, timedelta
from collections import defaultdict
from functools import wraps
import json

# Add to your imports section
import zipfile
from docxtpl import DocxTemplate
from io import BytesIO
from flask import send_file

#----------------------------------------------------------------------------------


#--------------------------------------------------------------
# --- Subscription Plan Definitions ---
PLANS = {
    'free_trial': {
        'name': 'Free Trial',
        'monthly_leads': 200,
        'monthly_emails': 5,
        'connected_accounts': 1,
        'cold_emails': 100,
        'document_generation': False,
        'trial_days': 14,
        'price': 0
    },
    'pilot': {
        'name': 'Pilot',
        'monthly_leads': 300,
        'monthly_emails': 300,
        'connected_accounts': 1,
        'cold_emails': 150,
        'document_generation': False,
        'trial_days': 0,
        'price': 19 
    },
    'starter': {
        'name': 'Starter',
        'monthly_leads': 500,
        'monthly_emails': 500,
        'connected_accounts': 1,
        'cold_emails': 200,
        'document_generation': False,
        'trial_days': 0,
        'price': 29
    },
    'professional': {
        'name': 'Professional',
        'monthly_leads': 2000,
        'monthly_emails': 2000,
        'connected_accounts': 3,
        'cold_emails': 1000,
        'document_generation': True,
        'trial_days': 0,
        'price': 79
    },
    'elite': {
        'name': 'Elite',
        'monthly_leads': 1000000,  # Very high number instead of infinite
        'monthly_emails': 1000000,
        'connected_accounts': 100,  # Very high number instead of infinite
        'cold_emails': 1000000,
        'document_generation': True,
        'trial_days': 0,
        'price': 199
    }
}



# --- Plan Rate Limiter Class (Modified for Profiles Table) ---
class PlanRateLimiter:
    def __init__(self, supabase_client):
        self.supabase = supabase_client
        self.local_cache = defaultdict(dict)

    def _reset_monthly_usage_if_needed(self, user_profile):
        """Reset monthly usage if it's a new month"""
        now = datetime.now(timezone.utc)
        reset_date = user_profile.get('usage_reset_date')
        
        if reset_date:
            if isinstance(reset_date, str):
                reset_date = datetime.fromisoformat(reset_date.replace('Z', '+00:00'))
            
            # Reset on the 1st of each month
            if now.month != reset_date.month or now.year != reset_date.year:
                update_data = {
                    'current_month_leads': 0,
                    'current_month_emails': 0,
                    'current_month_cold_emails': 0,
                    'usage_reset_date': now.isoformat()
                }
                
                self.supabase.table("profiles") \
                    .update(update_data) \
                    .eq("id", user_profile['id']) \
                    .execute()
                
                # Update local profile
                user_profile.update(update_data)
        
        return user_profile
    
    def get_user_plan(self, user_id):
        """Get user's current plan with trial status"""
        try:
            # Check cache first
            if user_id in self.local_cache and 'plan' in self.local_cache[user_id]:
                cached = self.local_cache[user_id]['plan']
                if datetime.now() - cached['fetched_at'] < timedelta(minutes=5):
                    return cached['data']
        
        # Get user's profile with plan info
            result = self.supabase.table("profiles") \
                .select("*") \
                .eq("id", user_id) \
                .single() \
                .execute()
        
            if result.data:
                profile = result.data
            
            # Reset monthly usage if needed
                profile = self._reset_monthly_usage_if_needed(profile)
            
                plan_name = profile.get('plan_name', 'starter').lower()
                subscription_status = profile.get('subscription_status', 'active')
            
            # Check if user is in trial period
                trial_ends_at = profile.get('trial_ends_at')
                trial_active = False
            
                if trial_ends_at:
                    trial_ends = datetime.fromisoformat(trial_ends_at.replace('Z', '+00:00'))
                    trial_active = datetime.now(timezone.utc) < trial_ends
            
                if trial_active:
                    # User is in trial - give them selected plan features
                    base_plan = PLANS.get(plan_name, PLANS['professional']).copy()
                    plan_data = {
                        'name': base_plan['name'] + ' (Trial)',
                        'monthly_leads': base_plan['monthly_leads'],
                        'monthly_emails': base_plan['monthly_emails'],
                        'connected_accounts': base_plan['connected_accounts'],
                        'cold_emails': base_plan['cold_emails'],
                        'document_generation': base_plan['document_generation'],
                        'is_trial': True,
                        'trial_days': 14,
                        'trial_ends_at': trial_ends_at,
                        'subscription_status': 'trial',
                        'plan_last_updated': profile.get('plan_last_updated')
                    }
                else:
                    # Regular plan - use values from profile (these match DB column names)
                    plan_data = {
                        'name': PLANS.get(plan_name, PLANS['starter']).get('name', 'Starter'),
                        'monthly_leads': profile.get('monthly_leads_limit', 200),
                        'monthly_emails': profile.get('monthly_emails_limit', 200),
                        'connected_accounts': profile.get('connected_accounts_limit', 1),
                        'cold_emails': profile.get('monthly_cold_emails_limit', 100),
                        'document_generation': profile.get('document_generation_enabled', False),
                        'is_trial': False,
                        'trial_days': 0,
                        'subscription_status': subscription_status,
                        'plan_last_updated': profile.get('plan_last_updated')
                    }
            
            # Add current usage from profile (these match DB column names)
                plan_data.update({
                    'current_leads': profile.get('current_month_leads', 0),
                    'current_emails': profile.get('current_month_emails', 0),
                    'current_cold_emails': profile.get('current_month_cold_emails', 0)
                })
            
            # Cache the result
                self.local_cache[user_id]['plan'] = {
                    'data': plan_data,
                    'fetched_at': datetime.now()
                }
            
                return plan_data
        
        # No profile found - default to starter
            default_plan = PLANS['starter'].copy()
            default_plan.update({
                'is_trial': False,
                'current_leads': 0,
                'current_emails': 0,
                'current_cold_emails': 0
            })
            return default_plan
        
        except Exception as e:
            app.logger.error(f"Error getting user plan: {str(e)}")
            # Fall back to starter plan
            fallback = PLANS['starter'].copy()
            fallback.update({
                'is_trial': False,
                'current_leads': 0,
                'current_emails': 0,
                'current_cold_emails': 0
            })
            return fallback
    
    def check_rate_limit(self, user_id, resource_type, amount=1):
        """
        Check if user has exceeded rate limit for a resource
        Returns: (allowed, remaining, message)
        """
        try:
            # First, get the latest plan data (ensures fresh usage counts)
            plan = self.get_user_plan(user_id)
            
            # Map resource types to plan limits
            resource_map = {
                'leads': ('monthly_leads', 'current_leads'),
                'emails': ('monthly_emails', 'current_emails'),
                'cold_emails': ('cold_emails', 'current_cold_emails'),
                'connected_accounts': ('connected_accounts', None)
            }
            
            if resource_type not in resource_map:
                return False, 0, f"Unknown resource type: {resource_type}"
            
            limit_key, current_key = resource_map[resource_type]
            plan_limit = plan.get(limit_key, 0)
            current_usage = plan.get(current_key, 0) if current_key else 0
            
            app.logger.info(f"Rate limit check for user {user_id}: {resource_type}")
            app.logger.info(f"  Plan limit: {plan_limit}, Current usage: {current_usage}, Requested: {amount}")
            
            # Check if adding amount would exceed limit
            if current_usage + amount > plan_limit:
                remaining = max(0, plan_limit - current_usage)
                message = f"{resource_type.replace('_', ' ').title()} limit exceeded. Plan limit: {plan_limit}, Used: {current_usage}"
                app.logger.warning(f"Rate limit exceeded: {message}")
                return False, remaining, message
            
            # If allowed, return success (actual increment happens elsewhere)
            remaining = plan_limit - current_usage
            return True, remaining, f"Limit: {plan_limit}, Used: {current_usage}, Remaining: {remaining}"
            
        except Exception as e:
            app.logger.error(f"Error checking rate limit: {str(e)}", exc_info=True)
            return False, 0, f"Error checking limits: {str(e)}"
    
    def _increment_usage(self, user_id, resource_type, amount=1):
        """Increment usage counter in database"""
        try:
            # Map resource types to column names
            column_map = {
                'leads': 'current_month_leads',
                'emails': 'current_month_emails',
                'cold_emails': 'current_month_cold_emails'
            }
        
            if resource_type not in column_map:
                return
        
            column = column_map[resource_type]
            
        # Use RPC function to increment - this is the most reliable
            try:
                self.supabase.rpc('increment_usage', {
                    'user_id': user_id,
                    'column_name': column,
                    'amount': amount
                }).execute()
            except Exception as rpc_error:
                # Fallback to direct update if RPC fails
                app.logger.warning(f"RPC increment failed, using direct update: {str(rpc_error)}")
                # Get current value first
                result = self.supabase.table("profiles") \
                    .select(column) \
                    .eq("id", user_id) \
                    .single() \
                    .execute()
                
                current_value = result.data.get(column, 0) if result.data else 0
                new_value = current_value + amount
            
            # Update the value
                self.supabase.table("profiles") \
                    .update({column: new_value}) \
                    .eq("id", user_id) \
                    .execute()
            
        except Exception as e:
            app.logger.error(f"Error incrementing usage: {str(e)}")
    
    def check_document_generation(self, user_id):
        """Check if user has document generation feature"""
        plan = self.get_user_plan(user_id)
        return plan.get('document_generation', False)
    
    def get_plan_info(self, user_id):
        """Get comprehensive plan information for display"""
        plan = self.get_user_plan(user_id)
        
        # Count connected email accounts
        connected_accounts = 0
        try:
            result = self.supabase.table("profiles") \
                .select("smtp_enc_password") \
                .eq("id", user_id) \
                .single() \
                .execute()
            
            if result.data and result.data.get('smtp_enc_password'):
                connected_accounts = 1
        except:
            pass
        
        return {
            'plan_name': plan['name'],
            'is_trial': plan.get('is_trial', False),
            'trial_days_left': self.get_trial_days_left(user_id) if plan.get('is_trial') else 0,
            'features': {
                'document_generation': plan.get('document_generation', False),
                'connected_accounts': {
                    'used': connected_accounts,
                    'limit': plan.get('connected_accounts', 1),
                    'allowed': connected_accounts < plan.get('connected_accounts', 1) or 
                               plan.get('connected_accounts', 1) >= 100  # Elite plan
                }
            },
            'usage': {
                'leads': {
                    'used': plan.get('current_leads', 0),
                    'limit': plan.get('monthly_leads', 500),
                    'remaining': max(0, plan.get('monthly_leads', 500) - plan.get('current_leads', 0))
                },
                'emails': {
                    'used': plan.get('current_emails', 0),
                    'limit': plan.get('monthly_emails', 500),
                    'remaining': max(0, plan.get('monthly_emails', 500) - plan.get('current_emails', 0))
                },
                'cold_emails': {
                    'used': plan.get('current_cold_emails', 0),
                    'limit': plan.get('cold_emails', 200),
                    'remaining': max(0, plan.get('cold_emails', 200) - plan.get('current_cold_emails', 0))
                }
            },
            'limits': {
                'monthly_leads': plan.get('monthly_leads', 500),
                'monthly_emails': plan.get('monthly_emails', 500),
                'cold_emails': plan.get('cold_emails', 200),
                'connected_accounts': plan.get('connected_accounts', 1)
            }
        }
    
    def get_trial_days_left(self, user_id):
        """Get remaining trial days"""
        try:
            result = self.supabase.table("profiles") \
                .select("trial_ends_at") \
                .eq("id", user_id) \
                .single() \
                .execute()
            
            if result.data and result.data.get('trial_ends_at'):
                trial_ends = datetime.fromisoformat(result.data['trial_ends_at'].replace('Z', '+00:00'))
                days_left = (trial_ends - datetime.now(timezone.utc)).days
                return max(0, days_left)
            
            return 0
            
        except Exception as e:
            app.logger.error(f"Error getting trial days: {str(e)}")
            return 0

    def update_user_plan(self, user_id, plan_name, start_trial=False):
        """Update user's plan in database"""
        try:
            if plan_name not in PLANS:
                return False, "Invalid plan name"
            
            plan_config = PLANS[plan_name]
            now = datetime.now(timezone.utc)
            
            update_data = {
                'plan_name': plan_name,
                'monthly_leads_limit': plan_config['monthly_leads'],
                'monthly_emails_limit': plan_config['monthly_emails'],
                'monthly_cold_emails_limit': plan_config['cold_emails'],
                'connected_accounts_limit': plan_config['connected_accounts'],
                'document_generation_enabled': plan_config['document_generation'],
                'plan_last_updated': now.isoformat()
            }
            
            if start_trial:
                trial_ends = now + timedelta(days=plan_config['trial_days'])
                update_data.update({
                    'trial_started_at': now.isoformat(),
                    'trial_ends_at': trial_ends.isoformat(),
                    'subscription_status': 'trial'
                })
            else:
                update_data.update({
                    'subscription_status': 'active',
                    'trial_started_at': None,
                    'trial_ends_at': None
                })
            
            # Update profile
            self.supabase.table("profiles") \
                .update(update_data) \
                .eq("id", user_id) \
                .execute()
            
            # Clear cache
            if user_id in self.local_cache:
                self.local_cache.pop(user_id, None)
            
            return True, f"Plan updated to {plan_config['name']}"
            
        except Exception as e:
            app.logger.error(f"Error updating user plan: {str(e)}")
            return False, str(e)
#----------------------------------------------------------------
    
@app.route("/signin2")
def signin():
    user_id = request.args.get("user_id", "")
    return render_template("signin2.html", user_id=user_id)
#--------------------------------------------------------------
app.register_blueprint(autopilot_bp, url_prefix="/autopilot")
app.register_blueprint(public_bp)

# --- Supabase setup ---
SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_ANON_KEY = os.environ["SUPABASE_ANON_KEY"]
SUPABASE_SERVICE_ROLE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
supabase: Client = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
SUPABASE_SERVICE: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
# Edge Function base URL *without* trailing slash or endpoint
EDGE_BASE_URL = os.environ.get("EDGE_BASE_URL", "").rstrip("/")
ENCRYPTION_KEY = os.environ["ENCRYPTION_KEY"].encode()  # 32-url-safe-base64 bytes
fernet = Fernet(ENCRYPTION_KEY)
# Retry configuration for calling the Edge Function
MAX_RETRIES = 5
RETRY_BACKOFF_BASE = 2

# Define follow-up sequence (days after initial contact)
FOLLOW_UP_SEQUENCE = [
    {"delay_days": 0, "name": "Immediate Follow-up"},
    {"delay_days": 1, "name": "Day 1 Follow-up"},
    {"delay_days": 3, "name": "Day 3 Follow-up"},
    {"delay_days": 7, "name": "Day 7 Follow-up"},
    {"delay_days": 14, "name": "Day 14 Follow-up"},
    {"delay_days": 30, "name": "Day 30 Follow-up"},
]

#----------------------------------------------------------------------------

def clean_placeholders(text):
    """Removes common placeholders like [Your Email], [Your Phone Number], etc."""
    if not text:
        return text
    
    # List of specific patterns to look for inside brackets
    placeholders = [
        r"\[Your Real Estate Agency\]",
        r"\[Your Email\]",
        r"\[Your Phone Number?\]",
        r"\[Your Name\]",
        r"\[Your Website\]"
    ]
    
    # Join patterns with OR operator
    pattern = "|".join(placeholders)
    
    # Remove the placeholder and any trailing whitespace it might have left
    cleaned_text = re.sub(pattern, "", text, flags=re.IGNORECASE)
    
    # Optional: Clean up triple newlines left behind by the removal
    cleaned_text = re.sub(r'\n{3,}', '\n\n', cleaned_text).strip()
    return cleaned_text

# ---------------------------------------------------------------------------
def call_edge(endpoint_path: str, payload: dict, return_response: bool = False):
    url = f"{EDGE_BASE_URL}{endpoint_path}"
    app.logger.info(f"🔗 call_edge → URL: {url}")
    app.logger.info(f"🔗 call_edge → Payload: {payload}")

    headers = {
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "apikey":        SUPABASE_SERVICE_ROLE_KEY,
        "Content-Type":  "application/json"
    }

    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.post(url, json=payload, headers=headers, timeout=120)
            app.logger.info(f"↩️  Response [{resp.status_code}]: {resp.text}")

            if resp.status_code == 200:
                if return_response:
                    return resp
                else:
                    return True
            elif resp.status_code == 429:
                wait = RETRY_BACKOFF_BASE ** attempt
                app.logger.warning(f"[{endpoint_path}] Rate‐limited, retry {attempt+1}/{MAX_RETRIES} after {wait}s")
                time.sleep(wait)
                continue
            else:
                app.logger.error(f"[{endpoint_path}] Failed ({resp.status_code}): {resp.text}")
                if return_response:
                    return resp
                else:
                    return False
        except requests.RequestException as e:
            wait = RETRY_BACKOFF_BASE ** attempt
            app.logger.error(f"[{endpoint_path}] Exception: {e}, retrying in {wait}s")
            time.sleep(wait)
    app.logger.error(f"[{endpoint_path}] Exceeded max retries.")
    if return_response:
        return None
    else:
        return False

# ── Routes ──
#-----------------------------------------------


# Add this near the top of your app.py after creating the Flask app
@app.template_filter('format_date')
def format_date_filter(value):
    if not value:
        return ""
    try:
        # Try to parse the date string
        date_obj = datetime.fromisoformat(value.replace('Z', '+00:00'))
        return date_obj.strftime("%b %d, %Y %I:%M %p")
    except:
        return value


#------------------------------------------------------------------

# Initialize rate limiter
rate_limiter = PlanRateLimiter(supabase)

def _require_user():
    uid = request.args.get("user_id") or request.form.get("user_id") or request.json.get("user_id")
    if not uid:
        abort(401, "Missing user_id")
    return uid

 
# ── Plan-aware Rate Limit Decorators ──
def check_plan_limit(resource_type, amount=1):
   # """Decorator to check plan-based rate limits"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user_id = _require_user()
            
            allowed, remaining, message = rate_limiter.check_rate_limit(user_id, resource_type, amount)
            
            if not allowed:
                return jsonify({
                    "error": "Plan limit exceeded",
                    "message": message,
                    "remaining": remaining,
                    "resource": resource_type
                }), 429
            
            # Add rate limit info to response headers
            response = make_response(f(*args, **kwargs))
            response.headers['X-RateLimit-Remaining'] = str(remaining)
            response.headers['X-RateLimit-Resource'] = resource_type
            
            return response
        return decorated_function
    return decorator


def require_feature(feature_name):
   # """Decorator to check if user has specific feature"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user_id = _require_user()
            
            if feature_name == 'document_generation':
                if not rate_limiter.check_document_generation(user_id):
                    return jsonify({
                        "error": "Feature not available",
                        "message": f"{feature_name.replace('_', ' ').title()} is not available in your plan",
                        "upgrade_required": True
                    }), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

#-------------------------------------------------
from flask import url_for

def get_safe_relay_address(display_name):
    """Generates a safe email alias from display name"""
    if not display_name:
        return "replyzeai.inbound+unknown@gmail.com"
    # Remove special chars and spaces, convert to lowercase
    safe_name = re.sub(r'[^a-zA-Z0-9]', '', display_name).lower()
    return f"replyzeai.inbound+{safe_name}@gmail.com"
    
#-------------------------------------------------------

@app.route("/")
def home():
    """
    Just redirect to /dashboard, passing along user_id if any.
    """
    user_id = request.args.get("user_id", "")
    # Redirect to /dashboard?user_id=<...> (blank if none)
    return redirect(f"/dashboard?user_id={user_id}")


@app.route("/dashboard")
def dashboard():
    user_id = request.args.get("user_id", "").strip()
    
    if not user_id:
        return redirect(f"/app/signin2?redirect=/dashboard")
    
    # Default values
    name = "Guest"
    ai_enabled = False
    generate_leases = False
    show_reconnect = False
    needs_connection = False  # Flag for the popup
    
    # Initialize counts
    replies_sent = 0
    outreach_emails = 0
    kits_generated = 0
    time_saved = 0
    relay_address = "replyzeai.inbound+unknown@gmail.com"
    
    # Get plan info with proper structure
    plan_info = rate_limiter.get_plan_info(user_id)
    
    if user_id:
        try:
            # Get profile - Updated to include smtp_enc_password and email
            profile_resp = supabase.table("profiles") \
                .select("full_name, ai_enabled, email, generate_leases, plan_name, display_name, current_month_emails, monthly_emails_limit, current_month_cold_emails, monthly_cold_emails_limit, smtp_enc_password, forwarding_verified") \
                .eq("id", user_id) \
                .single() \
                .execute()
            
            if profile_resp.data:
                profile = profile_resp.data
                name = profile.get("display_name") or profile.get("full_name", "Guest")
                ai_enabled = profile.get("ai_enabled", False)
                display_name = profile.get("display_name")
                generate_leases = profile.get("generate_leases", False)
                forwarding_verified = profile.get("forwarding_verified", False)
                relay_address = get_safe_relay_address(display_name)
                monthly_emails_limit = profile.get("monthly_emails_limit")
                
                # Check connections
                has_smtp = bool(profile.get("smtp_enc_password"))
                has_gmail = False
                
                try:
                    gmail_result = supabase.table("gmail_tokens") \
                        .select("credentials") \
                        .eq("user_id", user_id) \
                        .execute()
                    has_gmail = len(gmail_result.data or []) > 0
                except:
                    pass
                
                # Update connection flag: valid if SMTP OR Gmail OR Forwarding is verified
                needs_connection = not (has_smtp or has_gmail or forwarding_verified)
                
                # Check Gmail expiration
                if has_gmail:
                    try:
                        token_rows = supabase.table("gmail_tokens") \
                            .select("credentials") \
                            .eq("user_id", user_id) \
                            .execute().data or []
                        
                        if token_rows:
                            creds_data = token_rows[0]["credentials"]
                            creds = Credentials(
                                token=creds_data["token"],
                                refresh_token=creds_data["refresh_token"],
                                token_uri=creds_data["token_uri"],
                                client_id=creds_data["client_id"],
                                client_secret=creds_data["client_secret"],
                                scopes=creds_data["scopes"],
                            )
                            show_reconnect = creds.expired
                    except Exception:
                        show_reconnect = True
            
                # --- Count replies sent ---
                replies_result = supabase.table("emails") \
                    .select("id", count="exact") \
                    .eq("user_id", user_id) \
                    .eq("status", "sent") \
                    .execute()
                replies_sent = replies_result.count or 0
    
                # Increment the monthly emails count if we found sent emails
                if replies_sent > 0:
                    now = datetime.now(timezone.utc)
                    first_of_month = datetime(now.year, now.month, 1, tzinfo=timezone.utc).isoformat()
        
                    current_month_sent_result = supabase.table("emails") \
                        .select("id", count="exact") \
                        .eq("user_id", user_id) \
                        .eq("status", "sent") \
                        .gte("sent_at", first_of_month) \
                        .execute()
        
                    current_month_sent = current_month_sent_result.count or 0
                    current_profile = supabase.table("profiles") \
                        .select("current_month_emails") \
                        .eq("id", user_id) \
                        .single() \
                        .execute()
        
                    current_emails_count = current_profile.data.get("current_month_emails", 0) if current_profile.data else 0
        
                    if current_month_sent > current_emails_count:
                        increment_amount = current_month_sent - current_emails_count
                        if increment_amount > 0:
                            rate_limiter._increment_usage(user_id, 'emails', increment_amount)
            
                # Count outreach emails
                leads_result = supabase.table("leads") \
                    .select("id") \
                    .eq("user_id", user_id) \
                    .execute()
            
                lead_ids = [lead["id"] for lead in (leads_result.data or [])]
                if lead_ids:
                    outreach_result = supabase.table("lead_follow_ups") \
                        .select("id", count="exact") \
                        .in_("lead_id", lead_ids) \
                        .eq("status", "sent") \
                        .execute()
                    outreach_emails = outreach_result.count or 0
            
                # Count kits generated
                kits_result = supabase.table("transactions") \
                    .select("id", count="exact") \
                    .eq("user_id", user_id) \
                    .eq("kit_generated", True) \
                    .execute()
                kits_generated = kits_result.count or 0
            
                # Calculate time saved
                time_saved = replies_sent * 5.5 + kits_generated * 15
            
        except Exception as e:
            app.logger.error(f"Error loading dashboard: {str(e)}")
    
    # Build plan info for template
    plan_display_info = {
        'plan_name': plan_info.get('plan_name', 'Starter'),
        'is_trial': plan_info.get('is_trial', False),
        'trial_days_left': rate_limiter.get_trial_days_left(user_id) if plan_info.get('is_trial') else 0,
        'features': {
            'document_generation': plan_info.get('document_generation', False),
            'ai_replies': True,
            'outreach_emails': True,
            'multiple_accounts': plan_info.get('connected_accounts', 1) > 1,
            'unlimited_kits': plan_info.get('document_generation', False)
        },
        'usage': {
            'replies': plan_info.get('current_month_emails', 0),
            'outreach': plan_info.get('current_cold_emails', 0),
            'kits': 0,
            'limits': {
                'monthly_emails': plan_info.get('monthly_emails', 500),
                'cold_emails': plan_info.get('cold_emails', 200),
                'connected_accounts': plan_info.get('connected_accounts', 1)
            }
        },
        'current_counts': {
            'replies_sent': replies_sent,
            'outreach_emails': outreach_emails,
            'kits_generated': kits_generated
        }
    }
    
    return render_template(
        "dashboard.html",
        user_id=user_id,
        name=name,
        replies_sent=replies_sent,
        outreach_emails=outreach_emails,
        kits_generated=kits_generated,
        time_saved=time_saved,
        plan_info=plan_display_info,
        ai_enabled=ai_enabled,
        generate_leases=generate_leases,
        show_reconnect=show_reconnect,
        needs_connection=needs_connection,  # Pass the connection flag
        revenue=0,
        revenue_change=0,
        emails_sent=replies_sent,
        relay_address=relay_address,        # NEW
        forwarding_verified=forwarding_verified,
        monthly_emails_limit=monthly_emails_limit
    )
    #-------------------------------------------------------------------------------------------------------------------------------
@app.route("/api/check_relay_status")
def check_relay_status():
    user_id = request.args.get("user_id")
    if not user_id:
        return jsonify({"error": "Missing user_id"}), 400

    # In a real scenario, you might trigger a "Test Email" here to verify the loop.
    # For now, we fetch the DB status.
    result = supabase.table("profiles") \
        .select("forwarding_verified, forwarding_verified_at") \
        .eq("id", user_id) \
        .single() \
        .execute()
    
    if result.data:
        return jsonify({
            "verified": result.data.get("forwarding_verified", False),
            "verified_at": result.data.get("forwarding_verified_at")
        })
    return jsonify({"verified": False}), 200
    
    #--------------------------------------------------------------------------------------------------------------
@app.route("/dashboard/leads_funnel")
def dashboard_leads():
    user_id = _require_user()
    return render_template("partials/leads_funnel.html", user_id=user_id)

# Fix for the search error - update the leads_list function
@app.route("/dashboard/leads/list")
def leads_list():
    user_id = _require_user()
    filter_type = request.args.get("filter", "all")
    search_query = request.args.get("q", "")
    
    # Build query based on filters
    query = supabase.table("leads").select("*").eq("user_id", user_id)
    
    if filter_type != "all":
        query = query.eq("status", filter_type)
    
    # Execute query first to get all results
    try:
        result = query.execute()
        leads = result.data or []
    except Exception as e:
        app.logger.error(f"Error fetching leads: {str(e)}")
        leads = []
    
    # Apply search filter in Python
    if search_query:
        search_lower = search_query.lower()
        leads = [lead for lead in leads if 
                (lead.get("first_name", "").lower().find(search_lower) != -1 or
                 lead.get("last_name", "").lower().find(search_lower) != -1 or
                 lead.get("email", "").lower().find(search_lower) != -1 or
                 lead.get("brokerage", "").lower().find(search_lower) != -1)]
    
    # Calculate funnel counts
    counts = {
        "new": 0,
        "contacted": 0,
        "proposal": 0,
        "closed": 0
    }
    
    try:
        # Get counts for each status
        for status in counts.keys():
            count_result = supabase.table("leads").select("id", count="exact").eq("user_id", user_id).eq("status", status).execute()
            counts[status] = count_result.count or 0
    except Exception as e:
        app.logger.error(f"Error counting leads by status: {str(e)}")
    
    return render_template("partials/leads_list.html", leads=leads, counts=counts, user_id=user_id)


@app.route("/dashboard/leads/search")
def search_leads():
    # Reuse the leads_list function but with search parameters
    return leads_list()

@app.route("/dashboard/leads/<lead_id>/view")
def view_lead(lead_id):
    user_id = _require_user()
    
    try:
        # Get lead details
        lead = supabase.table("leads").select("*").eq("id", lead_id).eq("user_id", user_id).single().execute().data
        
        # Get follow-up history
        follow_ups = supabase.table("lead_follow_ups").select("*").eq("lead_id", lead_id).order("scheduled_at").execute().data or []
        
        return render_template("partials/lead_detail.html", lead=lead, follow_ups=follow_ups, user_id=user_id)
    except Exception as e:
        app.logger.error(f"Error fetching lead details: {str(e)}")
        return "<div class='error'>Error loading lead details: Missing required database columns</div>", 500

@app.route("/dashboard/leads/<lead_id>/update-status", methods=["POST"])
def update_lead_status(lead_id):
    user_id = _require_user()
    new_status = request.form.get("status")
    
    if not new_status:
        return jsonify({"error": "Status is required"}), 400
    
    try:
        # Update lead status
        supabase.table("leads").update({
            "status": new_status,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }).eq("id", lead_id).eq("user_id", user_id).execute()
        
        return "", 204
    except Exception as e:
        app.logger.error(f"Error updating lead status: {str(e)}")
        return jsonify({"error": "Failed to update status"}), 500

# Fix for the lead notes error - update the add_lead_note function
@app.route("/dashboard/leads/<lead_id>/add-note", methods=["POST"])
def add_lead_note(lead_id):
    user_id = _require_user()
    note_content = request.form.get("note")
    
    if not note_content:
        return jsonify({"error": "Note content is required"}), 400
    
    try:
        # First verify the lead exists and belongs to this user
        lead_check = supabase.table("leads").select("id").eq("id", lead_id).eq("user_id", user_id).execute()
        if not lead_check.data:
            return jsonify({"error": "Lead not found or access denied"}), 404
        
        # Add note to lead
        result = supabase.table("lead_notes").insert({
            "lead_id": lead_id,
            "user_id": user_id,
            "content": note_content,
            "created_at": datetime.now(timezone.utc).isoformat()
        }).execute()
        
        # Check if insertion was successful
        if not result.data:
            app.logger.error(f"Note insertion failed: {result}")
            return jsonify({"error": "Failed to add note - no data returned"}), 500
            
        return "", 204
    except Exception as e:
        app.logger.error(f"Error adding lead note: {str(e)}", exc_info=True)
        
        # Check if it's a specific API error
        error_msg = str(e)
        if "foreign key constraint" in error_msg.lower():
            return jsonify({"error": "Invalid lead ID"}), 400
        elif "null value" in error_msg.lower():
            return jsonify({"error": "Missing required fields"}), 400
            
        return jsonify({"error": "Failed to add note"}), 500


@app.route("/dashboard/leads/export")
def export_leads():
    user_id = _require_user()
    filter_type = request.args.get("filter", "all")
    
    try:
        # Build query
        query = supabase.table("leads").select("*").eq("user_id", user_id)
        
        if filter_type != "all":
            query = query.eq("status", filter_type)
        
        leads = query.execute().data or []
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(["First Name", "Last Name", "Email", "Brokerage", "Service", "City", "Status", "Last Contact"])
        
        # Write data
        for lead in leads:
            writer.writerow([
                lead.get("first_name", ""),
                lead.get("last_name", ""),
                lead.get("email", ""),
                lead.get("brokerage", ""),
                lead.get("service", ""),
                lead.get("city", ""),
                lead.get("status", "new"),
                lead.get("last_contacted_at", "")
            ])
        
        # Prepare response
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response.headers["Content-type"] = "text/csv"
        
        return response
    except Exception as e:
        app.logger.error(f"Error exporting leads: {str(e)}")
        return jsonify({"error": "Failed to export leads"}), 500
#------------------------------------------------------------------------------------------------------------
@app.route("/dashboard/new_transaction")
def dashboard_new_transaction():
    user_id = request.args.get("user_id") or abort(401)
    return render_template("partials/new_transaction.html", user_id=user_id)
  
@app.route("/dashboard/responded_emails")
def dashboard_responded_emails():
    user_id = request.args.get("user_id") or abort(401)
    # Select emails for this user that were sent/drafted and that have an original_content field
    try:
        emails = (
            supabase.table("emails")
                    .select("id, sender_email, subject, original_content, status, sent_at")
                    .eq("user_id", user_id)
                    .in_("status", ["sent","drafted"])   # treat drafted as 'responded' if you want
                    .order("sent_at", desc=True)
                    .execute()
                    .data
            or []
        )
    except Exception:
        app.logger.exception("failed to load responded emails")
        emails = []

    return render_template("partials/responded_emails.html", emails=emails, user_id=user_id)


@app.route("/dashboard/email/<email_id>/")
def dashboard_email_view(email_id):
    """Return a small partial showing full original_content — HTMX call for modal."""
    try:
        rec = supabase.table("emails").select("*").eq("id", email_id).single().execute().data
    except Exception:
        rec = None

    if not rec:
        return "<div class='chart-container'>Email not found.</div>", 404

    return render_template("partials/email_modal.html", email=rec)


@app.route("/dashboard/analytics")
def dashboard_analytics():
    user_id = _require_user()
    return render_template("partials/analytics.html", user_id=user_id)

@app.route("/dashboard/users")
def dashboard_users():
    user_id = _require_user()
    users = supabase.table("profiles").select("id, full_name, email").execute().data or []
    return render_template("partials/users.html", users=users)

@app.route("/dashboard/billing")
def dashboard_billing():
    user_id = _require_user()
    return render_template("partials/billing.html", user_id=user_id)

# Update the settings route to remove SMTP references
@app.route("/dashboard/settings", methods=["GET", "POST"])
def dashboard_settings():
    user_id = _require_user()

    if request.method == "POST":
        section = request.form.get("section")
        if section == "profile":
            new_display_name = request.form.get("display_name", "").strip()
            new_signature = request.form.get("signature", "").strip()
            supabase.table("profiles").update({
                "display_name": new_display_name,
                "signature": new_signature
            }).eq("id", user_id).execute()

    # Fetch profile & flags
    profile_resp = supabase.table("profiles") \
                           .select("display_name, signature, ai_enabled") \
                           .eq("id", user_id) \
                           .single() \
                           .execute()
    
    profile = profile_resp.data or {
        "display_name": "",
        "signature": "",
        "ai_enabled": False
    }

    # Check Gmail connection status
    gmail_connected = False
    show_reconnect = False
    
    try:
        service = get_gmail_service(user_id)
        if service:
            gmail_connected = True
            # Check if token needs refresh by trying a simple operation
            # If it fails, show reconnect button
            try:
                # Simple check to see if service works
                service.users().getProfile(userId='me').execute()
            except Exception:
                show_reconnect = True
    except Exception:
        app.logger.warning(f"settings: could not check Gmail token for {user_id}")

    # Render template
    return render_template(
        "partials/settings.html",
        profile=profile,
        user_id=user_id,
        gmail_connected=gmail_connected,
        show_reconnect=show_reconnect
    )



# Main dashboard with appointments
@app.route("/dashboard/home")
def dashboard_home():
    user_id = request.args.get("user_id")
    
    # Fetch today's appointments
    today = datetime.now().date()
    appointments_today = supabase.table("appointments") \
        .select("*, properties(address)") \
        .eq("user_id", user_id) \
        .gte("appointment_time", today.isoformat()) \
        .lt("appointment_time", (today + timedelta(days=1)).isoformat()) \
        .eq("status", "scheduled") \
        .execute().data
    
    # Format appointments for display
    formatted_appointments = []
    for apt in appointments_today:
        apt_time = datetime.fromisoformat(apt["appointment_time"])
        formatted_appointments.append({
            "id": apt["id"],
            "time": apt_time.strftime("%I:%M"),
            "period": apt_time.strftime("%p"),
            "client_email": apt["client_email"],
            "property_address": apt["properties"]["address"] if apt.get("properties") else "N/A"
        })
    
    # Get conversation stats
    conversations = supabase.table("conversations") \
        .select("*") \
        .eq("user_id", user_id) \
        .neq("conversation_stage", "completed") \
        .execute().data
    
    active_conversations = len(conversations)
    
    # Get properties
    properties = supabase.table("properties") \
        .select("*") \
        .eq("user_id", user_id) \
        .limit(5) \
        .execute().data
    
    # Get profile for email limits
    profile = supabase.table("profiles") \
        .select("*") \
        .eq("id", user_id) \
        .single() \
        .execute().data
    
    return render_template(
        "partials/home.html",
        user_id=user_id,
        name=profile.get("full_name", "Agent"),
        appointments_today=formatted_appointments,
        active_conversations=active_conversations,
        current_month_emails=profile.get("current_month_emails", 0),
        monthly_emails_limit=profile.get("monthly_emails_limit", 500),
        total_properties=len(properties),
        available_properties=len([p for p in properties if p["availability_status"] == "available"]),
        properties=properties[:3],
        current_date=datetime.now().strftime("%A, %B %d, %Y"),
        recent_activity=[],  # Add activity feed logic
        conversion_rate=0,  # Calculate from your data
        upcoming_appointments=len(appointments_today)
    )

# Properties page
@app.route("/dashboard/properties")
def properties_page():
    user_id = request.args.get("user_id")
    
    properties = supabase.table("properties") \
        .select("*") \
        .eq("user_id", user_id) \
        .execute().data
    
    return render_template(
        "partials/properties_page.html",
        user_id=user_id,
        properties=properties
    )




#-------------------------------------------------------------------------------------------------
# Add these routes to your app.py

from datetime import datetime, timedelta, timezone
from flask import request, render_template, jsonify

# ================================
# PROPERTIES ROUTES
# ================================

@app.route("/dashboard/properties/add", methods=["GET", "POST"])
def property_add():
    """Add new property form and handler"""
    user_id = request.args.get("user_id") or request.form.get("user_id")
    
    if request.method == "POST":
        try:
            # Get form data
            property_data = {
                "user_id": user_id,
                "address": request.form.get("address"),
                "property_type": request.form.get("property_type"),
                "bedrooms": int(request.form.get("bedrooms")) if request.form.get("bedrooms") else None,
                "bathrooms": float(request.form.get("bathrooms")) if request.form.get("bathrooms") else None,
                "square_feet": int(request.form.get("square_feet")) if request.form.get("square_feet") else None,
                "price": float(request.form.get("price")) if request.form.get("price") else None,
                "description": request.form.get("description"),
                "amenities": request.form.get("amenities"),
                "availability_status": "available",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Insert into database
            result = supabase.table("properties").insert(property_data).execute()
            
            if result.data:
                # Return success and trigger refresh
                return """
                <script>
                    window.location.reload();
                </script>
                """
            else:
                return jsonify({"error": "Failed to add property"}), 500
                
        except Exception as e:
            print(f"Error adding property: {e}")
            return jsonify({"error": str(e)}), 500
    
    # GET request - show form
    return render_template("partials/property_form.html", user_id=user_id, mode="add")


@app.route("/dashboard/properties/<property_id>/edit", methods=["GET", "POST"])
def property_edit(property_id):
    """Edit existing property"""
    user_id = request.args.get("user_id") or request.form.get("user_id")
    
    if request.method == "POST":
        try:
            # Update property
            update_data = {
                "address": request.form.get("address"),
                "property_type": request.form.get("property_type"),
                "bedrooms": int(request.form.get("bedrooms")) if request.form.get("bedrooms") else None,
                "bathrooms": float(request.form.get("bathrooms")) if request.form.get("bathrooms") else None,
                "square_feet": int(request.form.get("square_feet")) if request.form.get("square_feet") else None,
                "price": float(request.form.get("price")) if request.form.get("price") else None,
                "description": request.form.get("description"),
                "amenities": request.form.get("amenities"),
                "availability_status": request.form.get("availability_status", "available"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            result = supabase.table("properties") \
                .update(update_data) \
                .eq("id", property_id) \
                .eq("user_id", user_id) \
                .execute()
            
            return """
            <script>
                window.location.reload();
            </script>
            """
            
        except Exception as e:
            print(f"Error updating property: {e}")
            return jsonify({"error": str(e)}), 500
    
    # GET request - show form with existing data
    property_data = supabase.table("properties") \
        .select("*") \
        .eq("id", property_id) \
        .eq("user_id", user_id) \
        .single() \
        .execute().data
    
    return render_template(
        "partials/property_form.html",
        user_id=user_id,
        mode="edit",
        property=property_data
    )


@app.route("/dashboard/properties/<property_id>/details")
def property_view(property_id):
    """View property details"""
    user_id = request.args.get("user_id")
    
    # Get property with related data
    property_data = supabase.table("properties") \
        .select("*") \
        .eq("id", property_id) \
        .eq("user_id", user_id) \
        .single() \
        .execute().data
    
    # Get conversations/inquiries for this property
    conversations = supabase.table("conversations") \
        .select("*") \
        .eq("property_id", property_id) \
        .execute().data or []
    
    # Get appointments for this property
    appointments = supabase.table("appointments") \
        .select("*") \
        .eq("property_id", property_id) \
        .order("appointment_time", desc=False) \
        .execute().data or []
    
    return render_template(
        "partials/property_details.html",
        user_id=user_id,
        property=property_data,
        conversations=conversations,
        appointments=appointments
    )


@app.route("/dashboard/properties/<property_id>/delete", methods=["POST"])
def property_delete(property_id):
    """Delete a property"""
    user_id = request.form.get("user_id")
    
    try:
        supabase.table("properties") \
            .delete() \
            .eq("id", property_id) \
            .eq("user_id", user_id) \
            .execute()
        
        return """
        <script>
            window.location.reload();
        </script>
        """
    except Exception as e:
        print(f"Error deleting property: {e}")
        return jsonify({"error": str(e)}), 500


# ================================
# APPOINTMENTS ROUTES
# ================================

@app.route("/dashboard/appointments/new", methods=["GET", "POST"])
def appointment_new():
    """Create new appointment"""
    user_id = request.args.get("user_id") or request.form.get("user_id")
    conversation_id = request.args.get("conversation_id")
    
    if request.method == "POST":
        try:
            # Parse appointment date and time
            date_str = request.form.get("date")
            time_str = request.form.get("time")
            
            # Combine date and time
            appointment_datetime = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
            
            appointment_data = {
                "user_id": user_id,
                "client_email": request.form.get("client_email"),
                "client_name": request.form.get("client_name"),
                "property_id": request.form.get("property_id") or None,
                "conversation_id": request.form.get("conversation_id") or None,
                "appointment_time": appointment_datetime.isoformat(),
                "duration_minutes": 30,
                "status": "scheduled",
                "notes": request.form.get("notes"),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            result = supabase.table("appointments").insert(appointment_data).execute()
            
            # If there's a conversation, update it
            if conversation_id:
                supabase.table("conversations") \
                    .update({
                        "conversation_stage": "appointment_booked",
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }) \
                    .eq("id", conversation_id) \
                    .execute()
            
            return """
            <script>
                window.location.reload();
            </script>
            """
            
        except Exception as e:
            print(f"Error creating appointment: {e}")
            return jsonify({"error": str(e)}), 500
    
    # GET request - show form
    # Get user's properties for dropdown
    properties = supabase.table("properties") \
        .select("id, address") \
        .eq("user_id", user_id) \
        .eq("availability_status", "available") \
        .execute().data or []
    
    # If conversation_id provided, get conversation details
    conversation = None
    if conversation_id:
        conversation = supabase.table("conversations") \
            .select("*, properties(address)") \
            .eq("id", conversation_id) \
            .single() \
            .execute().data
    
    return render_template(
        "partials/appointment_form.html",
        user_id=user_id,
        mode="new",
        properties=properties,
        conversation=conversation
    )


@app.route("/dashboard/appointments/<appointment_id>")
def appointment_view(appointment_id):
    """View appointment details"""
    user_id = request.args.get("user_id")
    
    appointment = supabase.table("appointments") \
        .select("*, properties(address), conversations(*)") \
        .eq("id", appointment_id) \
        .eq("user_id", user_id) \
        .single() \
        .execute().data
    
    return render_template(
        "partials/appointment_details.html",
        user_id=user_id,
        appointment=appointment
    )


@app.route("/dashboard/appointments/<appointment_id>/complete", methods=["POST"])
def appointment_complee(appointment_id):
    """Mark appointment as complete"""
    user_id = request.form.get("user_id")
    
    try:
        # Update appointment status
        supabase.table("appointments") \
            .update({
                "status": "completed",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }) \
            .eq("id", appointment_id) \
            .eq("user_id", user_id) \
            .execute()
        
        # Get the appointment to update conversation
        apt = supabase.table("appointments") \
            .select("conversation_id") \
            .eq("id", appointment_id) \
            .single() \
            .execute().data
        
        # Update conversation to completed if exists
        if apt and apt.get("conversation_id"):
            supabase.table("conversations") \
                .update({
                    "conversation_stage": "completed",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }) \
                .eq("id", apt["conversation_id"]) \
                .execute()
        
        return """
        <script>
            window.location.reload();
        </script>
        """
        
    except Exception as e:
        print(f"Error completing appointment: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/dashboard/appointments/<appointment_id>/cancel", methods=["POST"])
def appointment_cancel(appointment_id):
    """Cancel an appointment"""
    user_id = request.form.get("user_id")
    
    try:
        supabase.table("appointments") \
            .update({
                "status": "cancelled",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }) \
            .eq("id", appointment_id) \
            .eq("user_id", user_id) \
            .execute()
        
        return """
        <script>
            window.location.reload();
        </script>
        """
        
    except Exception as e:
        print(f"Error cancelling appointment: {e}")
        return jsonify({"error": str(e)}), 500
        
        
#-------------------------------------------------------------------------------------------------        
# Add these routes to your app.py (in addition to the previous ones)

from datetime import datetime, timedelta, timezone

# ================================
# APPOINTMENTS PAGE (Full page with calendar/list)
# ================================

@app.route("/dashboard/appointments")
def appointments_page():
    """Full appointments page with calendar and list view"""
    user_id = request.args.get("user_id")
    period = request.args.get("period", "week")
    
    # Calculate date range based on period
    now = datetime.now()
    if period == "today":
        start_date = now.date()
        end_date = start_date + timedelta(days=1)
    elif period == "week":
        start_date = now.date()
        end_date = start_date + timedelta(days=7)
    else:  # month
        start_date = now.date()
        end_date = start_date + timedelta(days=30)
    
    # Fetch appointments
    appointments_raw = supabase.table("appointments") \
        .select("*, properties(address)") \
        .eq("user_id", user_id) \
        .gte("appointment_time", start_date.isoformat()) \
        .lt("appointment_time", end_date.isoformat()) \
        .order("appointment_time") \
        .execute().data or []
    
    # Generate week view data (for calendar)
    week_days = []
    for i in range(7):
        day_date = start_date + timedelta(days=i)
        day_appointments = []
        
        for apt in appointments_raw:
            apt_datetime = datetime.fromisoformat(apt["appointment_time"])
            if apt_datetime.date() == day_date:
                day_appointments.append({
                    "id": apt["id"],
                    "time": apt_datetime.strftime("%I:%M %p"),
                    "client_email": apt["client_email"],
                    "property_address": apt.get("properties", {}).get("address", "N/A") if apt.get("properties") else "N/A"
                })
        
        week_days.append({
            "name": day_date.strftime("%a"),
            "date": day_date.strftime("%d"),
            "is_today": day_date == now.date(),
            "appointments": day_appointments
        })
    
    # Format for list view
    formatted_appointments = []
    for apt in appointments_raw:
        apt_time = datetime.fromisoformat(apt["appointment_time"])
        formatted_appointments.append({
            "id": apt["id"],
            "date": apt_time.strftime("%B %d, %Y"),
            "time": apt_time.strftime("%I:%M %p"),
            "client_email": apt["client_email"],
            "client_name": apt.get("client_name"),
            "property_address": apt.get("properties", {}).get("address", "N/A") if apt.get("properties") else "N/A",
            "status": apt["status"]
        })
    
    return render_template(
        "partials/appointments_page.html",
        user_id=user_id,
        week_days=week_days,
        appointments=formatted_appointments
    )


# ================================
# CONVERSATIONS ROUTES
# ================================

# REPLACE your existing conversations_page route with this fixed version

@app.route("/dashboard/conversations")
def conversations_page():
    """Full conversations page with filtering"""
    user_id = request.args.get("user_id")
    
    conversations = supabase.table("conversations") \
        .select("*, properties(address)") \
        .eq("user_id", user_id) \
        .neq("conversation_stage", "completed") \
        .order("updated_at", desc=True) \
        .execute().data or []
    
    # Get stage counts
    stage_counts = {
        "initial_inquiry": 0,
        "awaiting_availability": 0,
        "appointment_booked": 0
    }
    
    for conv in conversations:
        stage = conv.get("conversation_stage")
        if stage in stage_counts:
            stage_counts[stage] += 1
    
    # Format conversations with last message AND appointment_id
    formatted_conversations = []
    for conv in conversations:
        # Get last email in this conversation
        last_email = supabase.table("emails") \
            .select("original_content") \
            .eq("sender_email", conv["client_email"]) \
            .eq("user_id", user_id) \
            .order("created_at", desc=True) \
            .limit(1) \
            .execute().data
        
        last_message = last_email[0]["original_content"][:150] if last_email else "No messages yet"
        
        # Count emails in this conversation
        email_count = supabase.table("emails") \
            .select("id", count="exact") \
            .eq("sender_email", conv["client_email"]) \
            .eq("user_id", user_id) \
            .execute().count or 1
        
        # Get appointment_id if this conversation has a booked appointment
        appointment_id = None
        if conv.get("conversation_stage") == "appointment_booked":
            apt_result = supabase.table("appointments") \
                .select("id") \
                .eq("conversation_id", conv["id"]) \
                .execute().data
            if apt_result:
                appointment_id = apt_result[0]["id"]
        
        formatted_conversations.append({
            **conv,
            "last_message_preview": last_message,
            "property_address": conv.get("properties", {}).get("address") if conv.get("properties") else None,
            "updated_at_formatted": format_time_ago(conv["updated_at"]),
            "email_count": email_count,
            "appointment_id": appointment_id  # ADD THIS
        })
    
    return render_template(
        "partials/conversations_page.html",
        user_id=user_id,
        conversations=formatted_conversations,
        total_conversations=len(conversations),
        awaiting_response=stage_counts["awaiting_availability"],
        stage_counts=stage_counts
    )


@app.route("/dashboard/conversations/<conversation_id>")
def conversation_view(conversation_id):
    """View single conversation with full email thread"""
    user_id = request.args.get("user_id")
    
    # Get conversation
    conversation = supabase.table("conversations") \
        .select("*, properties(address)") \
        .eq("id", conversation_id) \
        .eq("user_id", user_id) \
        .single() \
        .execute().data
    
    # Get all emails in this conversation
    emails = supabase.table("emails") \
        .select("*") \
        .eq("sender_email", conversation["client_email"]) \
        .eq("user_id", user_id) \
        .order("created_at", desc=False) \
        .execute().data or []
    
    # Get appointment if booked
    appointment = None
    if conversation.get("conversation_stage") == "appointment_booked":
        apt_data = supabase.table("appointments") \
            .select("*") \
            .eq("conversation_id", conversation_id) \
            .execute().data
        if apt_data:
            appointment = apt_data[0]
    
    return render_template(
        "partials/conversation_details.html",
        user_id=user_id,
        conversation=conversation,
        emails=emails,
        appointment=appointment
    )

@app.route("/dashboard/appointment/<conversation_id>/complete", methods=["POST"])
def appointment_complete(conversation_id):
    """Mark conversation as complete"""
    user_id = request.form.get("user_id")
    
    try:
        supabase.table("conversations") \
            .update({
                "conversation_stage": "completed",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }) \
            .eq("id", conversation_id) \
            .eq("user_id", user_id) \
            .execute()
        
        return """
        <script>
            window.location.reload();
        </script>
        """
    except Exception as e:
        print(f"Error completing conversation: {e}")
        return jsonify({"error": str(e)}), 500
        
        
@app.route("/dashboard/conversations/<conversation_id>/complete", methods=["POST"])
def conversation_complete(conversation_id):
    """Mark conversation as complete"""
    user_id = request.form.get("user_id")
    
    try:
        supabase.table("conversations") \
            .update({
                "conversation_stage": "completed",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }) \
            .eq("id", conversation_id) \
            .eq("user_id", user_id) \
            .execute()
        
        return """
        <script>
            window.location.reload();
        </script>
        """
    except Exception as e:
        print(f"Error completing conversation: {e}")
        return jsonify({"error": str(e)}), 500


# ================================
# HELPER FUNCTIONS
# ================================

def format_time_ago(timestamp_str):
    """Format timestamp as '5 minutes ago', '2 hours ago', etc."""
    if not timestamp_str:
        return "Unknown"
    
    try:
        timestamp = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
        now = datetime.now(timezone.utc)
        diff = now - timestamp
        
        seconds = diff.total_seconds()
        
        if seconds < 60:
            return "Just now"
        elif seconds < 3600:
            minutes = int(seconds / 60)
            return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
        elif seconds < 86400:
            hours = int(seconds / 3600)
            return f"{hours} hour{'s' if hours != 1 else ''} ago"
        else:
            days = int(seconds / 86400)
            return f"{days} day{'s' if days != 1 else ''} ago"
    except:
        return "Unknown"        
#-------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------

#----------------------------------------------------------------------

#------------------------------------------ 

@app.route("/complete_profile", methods=["GET", "POST"])
def complete_profile():
    user_id = request.args.get("user_id")
    if not user_id:
        return "Missing user_id", 401

    if request.method == "POST":
        display_name = request.form.get("display_name", "").strip()
        signature    = request.form.get("signature", "").strip()

        supabase.table("profiles") \
            .update({"display_name": display_name,
                     "signature": signature}) \
            .eq("id", user_id) \
            .execute()

        return redirect(f"/dashboard?user_id={user_id}")

    return render_template("complete_profile.html", user_id=user_id)

@app.route("/disconnect_gmail", methods=["POST"])
def disconnect_gmail():
    user_id = request.form.get("user_id")
    supabase.table("gmail_tokens").delete().eq("user_id", user_id).execute()
    return redirect(f"/dashboard?user_id={user_id}")

def _require_user():
    uid = request.args.get("user_id") or request.form.get("user_id")
    if not uid:
        abort(401, "Missing user_id")
    return uid

@app.route("/new_lease", methods=["GET"])
def new_lease_form():
    user_id = _require_user()
    return render_template("new_lease.html", user_id=user_id)

@app.route("/new_lease", methods=["POST"])
def new_lease_submit():
    user_id = _require_user()

    data = {
        "property_name":    request.form["propertyName"],
        "property_type":    request.form["propertyType"],
        "address":          request.form["address"],
        "suite":            request.form.get("suite",""),
        "square_feet":      request.form["squareFeet"],
        "tenant_name":      request.form["tenantName"],
        "tenant_type":      request.form["tenantType"],
        "lease_type":       request.form["leaseType"],
        "lease_term":       request.form["leaseTerm"],
        "start_date":       request.form["startDate"],
        "end_date":         request.form["endDate"],
        "base_rent":        request.form["baseRent"],
        "annual_increase":  request.form.get("annualIncrease",""),
        "security_deposit": request.form.get("securityDeposit",""),
        "parking_spaces":   request.form.get("parkingSpaces",""),
        "parking_fee":      request.form.get("parkingFee",""),
        "additional_terms": request.form.get("additionalTerms",""),
        "tenant_improvements": "Yes" if request.form.get("tenantImprovements") else "No",
        "renewal_option":      "Yes" if request.form.get("renewalOption") else "No",
        "exclusive_use":       "Yes" if request.form.get("exclusiveUse") else "No",
    }

    html_body = f"""
    <html><body>
      <h2>Lease Agreement</h2>
      <p><strong>Property:</strong> {data['property_name']} ({data['property_type'].title()})<br>
      <strong>Address:</strong> {data['address']} Suite {data['suite']}<br>
      <strong>Size:</strong> {data['square_feet']} sqft</p>

      <h3>Tenant</h3>
      <p>{data['tenant_name']} ({data['tenant_type'].title()})</p>

      <h3>Terms</h3>
      <p><strong>Type:</strong> {data['lease_type'].replace('-', '').title()}<br>
      <strong>Term:</strong> {data['lease_term']} months<br>
      <strong>Dates:</strong> {data['start_date']} → {data['end_date']}</p>

      <h3>Financials</h3>
      <p><strong>Base Rent:</strong> ${data['base_rent']} per sqft/yr<br>
      <strong>Annual Increase:</strong> {data['annual_increase']}%<br>
      <strong>Security Deposit:</strong> ${data['security_deposit']}<br>
      <strong>Parking:</strong> {data['parking_spaces']} spaces @ ${data['parking_fee']}/mo</p>

      <h3>Additional Terms</h3>
      <p>{data['additional_terms']}</p>
      <ul>
        <li>Tenant Improvements: {data['tenant_improvements']}</li>
        <li>Renewal Option: {data['renewal_option']}</li>
        <li>Exclusive Use Clause: {data['exclusive_use']}</li>
      </ul>
    </body></html>
    """

    tok = (supabase.table("gmail_tokens")
                .select("credentials")
                .eq("user_id", user_id)
                .limit(1)
                .execute()
                .data) or []
    if not tok:
        abort(400, "No Gmail token; reconnect Gmail first.")

    cd = tok[0]["credentials"]
    creds = Credentials(
        token=cd["token"],
        refresh_token=cd["refresh_token"],
        token_uri=cd["token_uri"],
        client_id=cd["client_id"],
        client_secret=cd["client_secret"],
        scopes=cd["scopes"],
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(GoogleRequest())

    service = build("gmail", "v1", credentials=creds, cache_discovery=False)

    mime = MIMEText(html_body, "html")
    mime["To"]      = ""
    mime["Subject"] = f"Draft Lease: {data['property_name']} → {data['tenant_name']}"
    raw = base64.urlsafe_b64encode(mime.as_bytes()).decode()
    draft = {"message": {"raw": raw}}
    created = service.users().drafts().create(userId="me", body=draft).execute()

    app.logger.info(f"Gmail Draft {created['id']} created for user {user_id}")

    return redirect(f"/dashboard?user_id={user_id}")

@app.route("/admin")
def admin():
    return render_template("admin.html")

@app.route("/api/admin/users")
def api_admin_users():
    users = supabase.table("profiles").select("*").execute().data or []
    today = date.today().isoformat()
    results = []
    for user in users:
        sent = supabase.table("emails") \
            .select("sent_at") \
            .eq("user_id", user["id"]) \
            .eq("status", "sent") \
            .execute().data or []
        count = len([e for e in sent if e["sent_at"] and e["sent_at"].startswith(today)])
        results.append({
            "id": user["id"],
            "name": user["full_name"],
            "email": user["email"],
            "enabled": user.get("ai_enabled", True),
            "emails_today": count
        })
    return jsonify(results)

@app.route("/api/admin/toggle_status", methods=["POST"])
def api_toggle_status():
    user_id = request.json.get("user_id")
    enable = request.json.get("enable", True)
    supabase.table("profiles").update({"ai_enabled": enable}).eq("id", user_id).execute()
    return jsonify({"success": True})

@app.route("/debug_env")
def debug_env():
    return {
        "GOOGLE_CLIENT_ID": os.environ.get("GOOGLE_CLIENT_ID"),
        "REDIRECT_URI": os.environ.get("REDIRECT_URI"),
        "EDGE_BASE_URL": os.environ.get("EDGE_BASE_URL")
    }

from datetime import datetime
@app.route("/process", methods=["GET"])
def trigger_process():
    token = request.args.get("token")
    if token != os.environ.get("PROCESS_SECRET_TOKEN"):
        app.logger.error("❌ Unauthorized process attempt")
        return jsonify({"error": "Unauthorized"}), 401

    app.logger.info("🚀 Starting optimized email processing...")

    # ── 0) DAILY RESET CHECK ──
    today_str = date.today().isoformat()
    rl_row = SUPABASE_SERVICE.table("rate_limit_reset") \
        .select("last_reset") \
        .eq("id", "global") \
        .single() \
        .execute().data or {}
    last_date = rl_row.get("last_reset", "")[:10]

    if last_date != today_str:
        app.logger.info("🔄 New day detected – clearing emails table")
        SUPABASE_SERVICE.table("emails") \
            .delete() \
            .neq("id", "00000000-0000-0000-0000-000000000000") \
            .execute()
        SUPABASE_SERVICE.table("rate_limit_reset") \
            .update({"last_reset": datetime.now(timezone.utc).isoformat()}) \
            .eq("id", "global") \
            .execute()

    BATCH_SIZE = 10
    all_processed, sent, failed = [], [], []

    # ── 1) TRANSITION PREPROCESSING → PROCESSING ──
    app.logger.info("🔄 Moving preprocessing emails to processing...")
    preprocessing = supabase.table("emails") \
        .select("id, user_id") \
        .eq("status", "preprocessing") \
        .limit(BATCH_SIZE) \
        .execute().data or []
    
    if preprocessing:
        preprocessing_ids = [row["id"] for row in preprocessing]
        app.logger.info(f"📥 Found {len(preprocessing_ids)} preprocessing emails")
        
        # Move them to processing
        supabase.table("emails") \
            .update({"status": "processing"}) \
            .in_("id", preprocessing_ids) \
            .execute()
        
        app.logger.info(f"✅ Moved {len(preprocessing_ids)} emails to processing")

    # ── 2) DIRECT PROCESSING LOOP FOR STUCK EMAILS ──
    # This is the NEW entrance you requested
    app.logger.info("🔍 Checking for emails stuck at 'processing' status...")
    stuck_processing = supabase.table("emails") \
        .select("id, user_id") \
        .eq("status", "processing") \
        .limit(BATCH_SIZE) \
        .execute().data or []
    
    if stuck_processing:
        app.logger.info(f"🔧 Found {len(stuck_processing)} emails at processing status - calling edge function directly")
        
        # Group by user
        stuck_by_user = defaultdict(list)
        for row in stuck_processing:
            user_id = row.get("user_id")
            if user_id:
                stuck_by_user[user_id].append(row["id"])
        
        # Process each user's emails
        for user_id, email_ids in stuck_by_user.items():
            app.logger.info(f"🤖 Calling Edge Function for {len(email_ids)} stuck emails from user {user_id}")
            
            # Check rate limit
            allowed, remaining, message = rate_limiter.check_rate_limit(user_id, 'emails', len(email_ids))
            
            if not allowed:
                app.logger.warning(f"⚠️ Rate limit exceeded for user {user_id}")
                supabase.table("emails") \
                    .update({"status": "rate_limited", "error_message": message}) \
                    .in_("id", email_ids).execute()
                failed.extend(email_ids)
                continue
            
            # Call Edge Function
            try:
                edge_result = call_edge("/functions/v1/clever-agent/generate-response", {"email_ids": email_ids})
                
                if edge_result:
                    app.logger.info(f"✅ Edge function succeeded for {len(email_ids)} stuck emails")
                    # Move to ready_to_send
                    supabase.table("emails") \
                        .update({"status": "ready_to_send", "processed_at": datetime.utcnow().isoformat()}) \
                        .in_("id", email_ids).execute()
                    all_processed.extend(email_ids)
                else:
                    app.logger.error(f"❌ Edge function failed for stuck emails: {email_ids}")
                    supabase.table("emails") \
                        .update({"status": "error", "error_message": "AI Generation Failed"}) \
                        .in_("id", email_ids).execute()
                    failed.extend(email_ids)
            except Exception as e:
                app.logger.error(f"❌ Exception processing stuck emails: {e}")
                supabase.table("emails") \
                    .update({"status": "error", "error_message": str(e)}) \
                    .in_("id", email_ids).execute()
                failed.extend(email_ids)

    # ── 3) SMALL DELAY FOR DB CONSISTENCY ──
    if all_processed:
        import time
        time.sleep(0.5)
        app.logger.info("⏳ Allowing database to sync...")

    # ── 4) FETCH AND SEND EMAILS (READY_TO_SEND) ──
    ready = (
        supabase.table("emails")
        .select("id, user_id, sender_email, recipient_email, processed_content, subject")
        .eq("status", "ready_to_send")
        .limit(BATCH_SIZE)
        .execute()
        .data or []
    )

    app.logger.info(f"📨 Found {len(ready)} emails ready to send")

    if ready:
        ready_by_user = defaultdict(list)
        for rec in ready:
            uid = rec.get("user_id")
            if uid:
                ready_by_user[uid].append(rec)

        for user_id, user_emails in ready_by_user.items():
            for rec in user_emails:
                em_id = rec["id"]
                try:
                    admin_email = os.environ.get("ADMIN_INBOUND_EMAIL")
                    smtp_password = os.environ.get("ADMIN_INBOUND_PASSWORD")
                    
                    # Get profile for signature
                    prof = supabase.table("profiles").select("display_name, signature").eq("id", user_id).single().execute().data

                    body_html = (rec.get("processed_content") or "").replace("\n", "<br>")
                    if prof and prof.get("signature"):
                        body_html += f"<br><br>{prof['signature']}"
                    if prof and prof.get("display_name"):
                        body_html = body_html.replace("[Your Name]", prof["display_name"])

                    send_email_smtp(
                        from_email=admin_email,
                        from_password=smtp_password,
                        to_email=rec["sender_email"],
                        subject=f"RE: {rec.get('subject', 'Update')}",
                        body=f"<html><body>{body_html}</body></html>",
                        smtp_host="smtp.gmail.com",
                        smtp_port=465
                    )

                    supabase.table("emails").update({
                        "status": "sent",
                        "sent_at": datetime.utcnow().isoformat()
                    }).eq("id", em_id).execute()
                    
                    rate_limiter._increment_usage(user_id, 'emails', 1)
                    sent.append(em_id)
                    app.logger.info(f"✅ Email {em_id} sent successfully")

                except Exception as e:
                    app.logger.error(f"❌ SMTP failed for {em_id}: {str(e)}")
                    supabase.table("emails").update({"status": "error", "error_message": str(e)}).eq("id", em_id).execute()
                    failed.append(em_id)

    # ── 5) RETURN SUMMARY ──
    response = {
        "preprocessing_moved": len(preprocessing) if preprocessing else 0,
        "stuck_processing_found": len(stuck_processing) if stuck_processing else 0,
        "processed_in_batch": len(all_processed),
        "sent": len(sent),
        "failed": len(failed)
    }
    
    app.logger.info(f"📊 Summary: {response}")
    return jsonify(response), 200
#---------------------------------------------------------------------------------------------------------------------------



#-----------------------------------------------------------------------------------------------------

@app.route("/transaction/<txn_id>/ready", methods=["POST"])
def mark_ready(txn_id):
    supabase.table("transactions").update({"ready_for_kit": True}).eq("id", txn_id).execute()
    return "", 204

@app.route("/autopilot/batch", methods=["POST"])
@check_plan_limit('kits', amount=1)  # Each kit counts as a lead
@require_feature('document_generation')
def batch_autopilot():
    user_id = _require_user()
    
    if not rate_limiter.check_document_generation(user_id):
        return jsonify({
            "error": "Document generation not available",
            "message": "Upgrade to Professional or Elite plan for document generation"
        }), 403
    
    txns = supabase.table("transactions").select("*").eq("ready_for_kit", True).eq("kit_generated", False).execute().data or []
    results = []
    for t in txns:
        payload = {
          "transaction_type": t["transaction_type"],
          "data": {
            "id": t["id"],
            "buyer": t["buyer"],
            "seller": t["seller"],
            "date": t["date"],
            "purchase_price": t["purchase_price"],
            "closing_date": t.get("closing_date"),
            "closing_location": t.get("closing_location")
          }
        }
        resp = requests.post(f"{os.environ.get('BASE_URL')}/autopilot/trigger", json=payload)
        results.append({"id": t["id"], "status": resp.status_code})
        if resp.ok:
            supabase.table("transactions").update({"kit_generated": True}).eq("id", t["id"]).execute()
    return jsonify(results), 200

@app.route("/dashboard/autopilot")
def dashboard_autopilot():
    user_id = request.args.get("user_id") or abort(401)
    txn_id  = request.args.get("txn_id")
    transactions = supabase.table("transactions").select("*").eq("user_id", user_id).execute().data or []
    current_txn = None
    if txn_id:
        resp = supabase.table("transactions").select("*").eq("id", txn_id).execute()
        current_txn = resp.data[0] if resp.data else None
    return render_template("partials/autopilot.html", user_id=user_id, transactions=transactions, current_transaction=current_txn)

@app.route("/transactions/new", methods=["POST"])
def create_transaction():
    import uuid
    import traceback

    user_id = request.args.get("user_id") or request.form.get("user_id")
    if not user_id:
        return jsonify({"status": "error", "message": "Missing user_id"}), 401

    new_id = str(uuid.uuid4())

    # 🔐 Validate required fields (lowercase unified names)
    required = ["buyer_name", "seller_name", "property_address", "agreement_date"]
    missing = [f for f in required if not request.form.get(f)]
    if missing:
        app.logger.warning(f"⚠️ Missing required fields: {missing}")
        return jsonify({
            "status": "error",
            "message": f"Missing required fields: {', '.join(missing)}"
        }), 400

    # ✅ All accepted lowercase fields from gamified form
    accepted_fields = [
        "transaction_type", "property_address", "city", "state", "name_of_property",
        "description_of_property", "square_feet", "legal_description",
        "apartment_address", "premises_description",

        "buyer_name", "buyer_address", "seller_name", "seller_address", "agency_name",

        "purchase_price", "deposit_amount", "agreement_date", "broker_name",
        "commission_amount", "brokerage_fee", "broker_payday",

        "closing_date", "occupy_property_date", "mortgage_amount", "mortgage_years",
        "interest_rate", "inspection_days", "possession_date",

        "rent_type", "agreed_rent", "maintenance_terms",

        "landlord_phone", "tenant_phone", "landlord_email", "tenant_email",

        "structure_age", "location", "county", "additional_explanations",

        "buyer_signature", "seller_signature", "time"
    ]

    # Build the payload, turning empty strings into None
    payload = {"id": new_id, "user_id": user_id}
    for field in accepted_fields:
        val = request.form.get(field)
        payload[field] = None if val == "" else val

    try:
        app.logger.info(f"🚀 Inserting transaction with ID {new_id}")
        app.logger.debug(f"Payload: {payload}")
        resp = supabase.table("transactions").insert(payload).execute()
        inserted = resp.data[0]
    except Exception as e:
        app.logger.error("❌ Transaction insert failed")
        app.logger.error(traceback.format_exc())
        return jsonify({
            "status": "error",
            "message": f"Insertion failed: {str(e)}"
        }), 500

    # ✅ Success response with htmx trigger
    feedback = (
        f'<div class="alert alert-success">🎉 Transaction <strong>{inserted["id"]}</strong> created.</div>'
        + '<script>htmx.trigger(document.querySelector(\'[hx-get*="/dashboard/autopilot"]\'), "click")</script>'
    )
    return feedback, 200

# Add this to your main app file (e.g., app.py)
# Add these imports at the top of your app.py
import re
import dns.resolver

# Add this route to your app.py


def extract_domain(email):
    """Extract domain from email address"""
    pattern = r'@([\w\.-]+)'
    match = re.search(pattern, email)
    if match:
        return match.group(1).lower()
    return None




#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

import csv
from io import TextIOWrapper
from openpyxl import load_workbook

@app.route("/import_leads", methods=["GET", "POST"])
@check_plan_limit('leads')
def import_leads():
    user_id = _require_user()
    
    if request.method == "GET":
        # Get plan info to show limits
        plan_info = rate_limiter.get_plan_info(user_id)
        return render_template("import_leads.html", 
                             user_id=user_id, 
                             plan_info=plan_info)
    
    # Handle POST request
    try:
        # Debug logging
        app.logger.info(f"Import leads request received: {request.files}")
        
        if 'file' not in request.files:
            app.logger.error("No file in request")
            return jsonify({"error": "No file uploaded"}), 400
        
        file = request.files['file']
        if file.filename == '':
            app.logger.error("Empty filename")
            return jsonify({"error": "No file selected"}), 400
        
        # Check file extension
        if file.filename.endswith('.csv'):
            # Process CSV file
            csv_file = TextIOWrapper(file, encoding='utf-8')
            reader = csv.DictReader(csv_file)
            rows = list(reader)
            app.logger.info(f"CSV columns: {reader.fieldnames}")
        elif file.filename.endswith(('.xlsx', '.xls')):
            # Process Excel file
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1] if cell.value]
            
            # Get data rows
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                rows.append(row_data)
            
            app.logger.info(f"Excel columns: {headers}")
        else:
            app.logger.error(f"Invalid file type: {file.filename}")
            return jsonify({"error": "Invalid file type. Please upload CSV or Excel."}), 400
        
        # Check if we have any rows
        if not rows:
            app.logger.error("No data rows found in file")
            return jsonify({"error": "No data found in file"}), 400
        
        # Define mapping of possible column names to standard fields
        column_mappings = {
            # Standard fields mapping
            'first_name': ['first_name', 'First Name', 'first name', 'First', 'FirstName', 'given_name', 'Given Name'],
            'last_name': ['last_name', 'Last Name', 'last name', 'Last', 'LastName', 'surname', 'Surname', 'family_name', 'Family Name'],
            'email': ['email', 'Email', 'EMAIL', 'e-mail', 'E-mail', 'Recipient', 'recipient', 'email_address', 'Email Address'],
            'city': ['city', 'City', 'CITY', 'location', 'Location', 'town', 'Town', 'suburb', 'Suburb'],
            'brokerage': ['brokerage', 'Brokerage', 'company', 'Company', 'firm', 'Firm', 'organization', 'Organization'],
            'service': ['service', 'Service', 'interest', 'Interest', 'service_type', 'Service Type', 'product', 'Product'],
            'status': ['status', 'Status', 'stage', 'Stage', 'lead_status', 'Lead Status'],
            'email_sent': ['email_sent', 'Email Sent', 'contact_date', 'Contact Date', 'date_contacted', 'Date Contacted', 'timestamp', 'Timestamp']
        }
        
        # Process each row
        success_count = 0
        error_count = 0
        imported_leads = []
        failed_rows = []
        
        for i, row in enumerate(rows):
            try:
                # Clean row data - convert all values to strings and strip whitespace
                cleaned_row = {}
                for key, value in row.items():
                    if value is None:
                        cleaned_row[key] = ''
                    else:
                        cleaned_row[str(key).strip()] = str(value).strip()
                
                # Map columns to standard fields
                lead_data = {
                    'user_id': user_id,
                    'created_at': datetime.utcnow().isoformat()
                }
                
                extra_data = {}
                
                # Try to map each column
                for original_key, value in cleaned_row.items():
                    matched = False
                    
                    # Check if this column maps to a standard field
                    for standard_field, possible_names in column_mappings.items():
                        if original_key in possible_names:
                            lead_data[standard_field] = value
                            matched = True
                            break
                    
                    # If not matched to a standard field, add to extra_data
                    if not matched and value:  # Only add non-empty values
                        extra_data[original_key] = value
                
                # Check for required fields (email is the only true requirement)
                if not lead_data.get('email'):
                    # Try to find email by looking at column content
                    email_found = False
                    for key, value in cleaned_row.items():
                        if '@' in value and '.' in value and ' ' not in value:
                            lead_data['email'] = value
                            email_found = True
                            # Remove from extra_data if it was added there
                            if key in extra_data:
                                del extra_data[key]
                            break
                    
                    if not email_found:
                        error_message = f"Row {i+1}: No valid email address found"
                        failed_rows.append({"row": i+1, "error": error_message, "data": cleaned_row})
                        error_count += 1
                        continue
                
                # Parse email_sent date if available
                if lead_data.get('email_sent'):
                    try:
                        email_sent_str = lead_data['email_sent']
                        # Try different date formats
                        date_formats = [
                            '%Y-%m-%d %H:%M:%S',
                            '%Y-%m-%d',
                            '%m/%d/%Y %H:%M:%S',
                            '%m/%d/%Y',
                            '%d/%m/%Y %H:%M:%S',
                            '%d/%m/%Y',
                            '%Y/%m/%d %H:%M:%S',
                            '%Y/%m/%d'
                        ]
                        
                        email_sent_date = None
                        for date_format in date_formats:
                            try:
                                email_sent_date = datetime.strptime(email_sent_str, date_format)
                                break
                            except ValueError:
                                continue
                        
                        if email_sent_date:
                            lead_data['email_sent'] = email_sent_date.isoformat()
                        else:
                            # If can't parse, use current time
                            lead_data['email_sent'] = datetime.utcnow().isoformat()
                    except Exception as e:
                        app.logger.warning(f"Error parsing date '{lead_data.get('email_sent')}': {e}")
                        lead_data['email_sent'] = datetime.utcnow().isoformat()
                else:
                    lead_data['email_sent'] = datetime.utcnow().isoformat()
                
                # Set default status if not provided
                if not lead_data.get('status'):
                    lead_data['status'] = 'contacted'
                
                # Add extra_data JSON if we have any extra fields
                if extra_data:
                    lead_data['extra_data'] = extra_data
                
                # Validate email format
                if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', lead_data['email']):
                    error_message = f"Row {i+1}: Invalid email format '{lead_data['email']}'"
                    failed_rows.append({"row": i+1, "error": error_message, "data": cleaned_row})
                    error_count += 1
                    continue
                
                # Check rate limit
                allowed, remaining, message = rate_limiter.check_rate_limit(user_id, 'leads', 1)
                if not allowed:
                    error_message = f"Row {i+1}: Rate limit exceeded - {message}"
                    failed_rows.append({"row": i+1, "error": error_message, "data": cleaned_row})
                    error_count += 1
                    continue
                
                # Insert lead
                response = supabase.table('leads').insert(lead_data).execute()
                
                if response.data:
                    success_count += 1
                    lead_id = response.data[0]['id']
                    imported_leads.append(lead_id)
                    
                    # Increment usage counter
                    rate_limiter._increment_usage(user_id, 'leads', 1)
                    
                    # Schedule follow-ups
                    try:
                        # Send immediate follow-up (step 0)
                        follow_up_content = generate_follow_up_content(lead_id, 0)
                        
                        if follow_up_content:
                            # Get lead details
                            lead = supabase.table('leads').select('*').eq('id', lead_id).single().execute().data
                            
                            # Send email using Gmail API
                            success, message = send_email_gmail(
                                user_id,
                                lead['email'],
                                "Follow-up from your inquiry",
                                follow_up_content
                            )
                            
                            if success:
                                # Create follow-up record
                                follow_up_data = {
                                    'lead_id': lead_id,
                                    'sequence_step': 0,
                                    'generated_content': follow_up_content,
                                    'status': 'sent',
                                    'sent_at': datetime.utcnow().isoformat()
                                }
                                supabase.table('lead_follow_ups').insert(follow_up_data).execute()
                            else:
                                app.logger.error(f"Failed to send immediate follow-up for lead {lead_id}: {message}")
                    
                    except Exception as e:
                        app.logger.error(f"Error sending immediate follow-up for lead {lead_id}: {str(e)}")
                    
                    # Schedule the rest of the follow-up sequence
                    email_sent = datetime.fromisoformat(lead_data['email_sent'].replace('Z', '+00:00'))
                    for step, seq in enumerate(FOLLOW_UP_SEQUENCE[1:], start=1):
                        scheduled_at = email_sent + timedelta(days=seq['delay_days'])
                        follow_up_data = {
                            'lead_id': lead_id,
                            'sequence_step': step,
                            'scheduled_at': scheduled_at.isoformat(),
                            'status': 'pending'
                        }
                        supabase.table('lead_follow_ups').insert(follow_up_data).execute()
                
                else:
                    error_message = f"Row {i+1}: Database insertion failed"
                    failed_rows.append({"row": i+1, "error": error_message, "data": cleaned_row})
                    error_count += 1
            
            except Exception as e:
                error_message = f"Row {i+1}: {str(e)}"
                failed_rows.append({"row": i+1, "error": error_message, "data": cleaned_row if 'cleaned_row' in locals() else {}})
                error_count += 1
                app.logger.error(f"Error processing row {i+1}: {e}", exc_info=True)
        
        # Log summary
        app.logger.info(f"Import completed: {success_count} succeeded, {error_count} failed")
        
        response_data = {
            "message": f"Import completed. {success_count} leads imported successfully, {error_count} failed.",
            "imported_count": success_count,
            "failed_count": error_count,
            "success": success_count > 0
        }
        
        # Add failed rows details if any
        if failed_rows:
            response_data["failed_rows"] = failed_rows[:10]  # Limit to first 10 for response
            response_data["total_failed_rows"] = len(failed_rows)
        
        return jsonify(response_data), 200 if success_count > 0 else 207  # 207 for partial success
    
    except Exception as e:
        app.logger.error(f"Error importing leads: {str(e)}", exc_info=True)
        return jsonify({"error": f"Failed to import leads: {str(e)}"}), 500
#------------------------------------------------------------------------------------------------------------------
def generate_follow_up_content(lead_id, sequence_step):
    """Generate follow-up content using AI with context of previous communications"""
    try:
        app.logger.info(f"Starting follow-up generation for lead {lead_id}, step {sequence_step}")
        
        # Get lead details
        lead_resp = supabase.table("leads").select("*").eq("id", lead_id).single().execute()
        if not lead_resp.data:
            app.logger.error(f"Lead {lead_id} not found")
            return None
            
        lead = lead_resp.data
        app.logger.info(f"Found lead: {lead['email']}")
        
        # Get previous emails from emails table
        previous_emails = supabase.table("emails") \
            .select("subject, original_content, processed_content, sent_at") \
            .eq("sender_email", lead["email"]) \
            .order("sent_at", desc=True) \
            .limit(5) \
            .execute().data or []
        
        # Get previous follow-ups from lead_follow_ups table
        previous_follow_ups = supabase.table("lead_follow_ups") \
            .select("generated_content, sent_at, sequence_step") \
            .eq("lead_id", lead_id) \
            .eq("status", "sent") \
            .lt("sequence_step", sequence_step) \
            .order("sent_at", desc=True) \
            .execute().data or []
        
        app.logger.info(f"Found {len(previous_emails)} previous emails and {len(previous_follow_ups)} previous follow-ups")
        
        # Build context for AI
        context = f"""
        Lead: {lead['first_name']} {lead['last_name']}
        Company: {lead['brokerage']}
        Service: {lead['service']}
        Location: {lead['city']}
        
        Previous communications:
        """
        
        # Add emails from emails table
        for i, email in enumerate(previous_emails):
            context += f"\nEmail {i+1} ({email.get('sent_at', '')}):\n"
            context += f"Subject: {email.get('subject', 'No subject')}\n"
            content = email.get('original_content') or email.get('processed_content', '')
            context += f"Content: {content[:200]}...\n" if len(content) > 200 else f"Content: {content}\n"
        
        # Add follow-ups from lead_follow_ups table
        for i, follow_up in enumerate(previous_follow_ups, start=len(previous_emails)+1):
            context += f"\nFollow-up {i} (Day {FOLLOW_UP_SEQUENCE[follow_up['sequence_step']]['delay_days']}, {follow_up.get('sent_at', '')}):\n"
            content = follow_up.get('generated_content', '')
            context += f"Content: {content[:200]}...\n" if len(content) > 200 else f"Content: {content}\n"
        
        if not previous_emails and not previous_follow_ups:
            context += "\nNo previous communications found. This is the first contact.\n"
        
        context += f"\n\nWrite a friendly, professional follow-up email for day {FOLLOW_UP_SEQUENCE[sequence_step]['delay_days']}."
        context += " Reference previous communications if relevant. Keep it concise and focused on providing value."
        
        app.logger.info(f"Built context for AI: {context[:500]}...")
        
        # Call your AI API
        payload = {
            "context": context,
            "type": "follow_up",
            "sequence_step": sequence_step,
            "lead_id": lead_id
        }
        
        app.logger.info(f"Calling edge function with payload: {payload}")
        
        # Use your existing Edge Function call pattern
        # Modify call_edge to return the response content instead of just success/failure
        response = call_edge("/functions/v1/generate-follow-up", payload, return_response=True)
        
        if response and response.status_code == 200:
            content = response.json().get("content")
            app.logger.info(f"Successfully generated follow-up for lead {lead_id}")
            return content
        else:
            app.logger.error(f"Failed to generate follow-up content for lead {lead_id}")
            return None
            
    except Exception as e:
        app.logger.error(f"Error generating follow-up content: {str(e)}", exc_info=True)
        return None
#-------------------------------------------------------------------------------------------------------------------------------------------------


# Update the process_follow_ups route to use Gmail API
@app.route("/process_follow_ups", methods=["GET"])
def process_follow_ups():
    # Check for secret token
    token = request.args.get("token")
    if token != os.environ.get("PROCESS_SECRET_TOKEN"):
        return jsonify({"error": "Unauthorized"}), 401
    
    try:
        # Get due follow-ups
        now = datetime.now(timezone.utc).isoformat()
        due_follow_ups = supabase.table("lead_follow_ups") \
            .select("*, leads(*)") \
            .lte("scheduled_at", now) \
            .eq("status", "pending") \
            .execute().data
        
        results = {"processed": [], "failed": []}
        
        for follow_up in due_follow_ups:
            try:
                # Generate content using AI
                content = generate_follow_up_content(follow_up["lead_id"], follow_up["sequence_step"])
                if content:
                    # Get user_id from lead
                    user_id = follow_up["leads"]["user_id"]
                    lead_email = follow_up["leads"]["email"]
                    
                    # Send using Gmail API
                    subject = f"Follow-up: {follow_up['leads'].get('first_name', '')} {follow_up['leads'].get('last_name', '')}"
                    success, message = send_email_gmail(
                        user_id,
                        lead_email,
                        subject,
                        content
                    )
                    
                    if success:
                        # Update status
                        supabase.table("lead_follow_ups") \
                            .update({
                                "status": "sent", 
                                "generated_content": content,
                                "sent_at": now
                            }) \
                            .eq("id", follow_up["id"]) \
                            .execute()
                        results["processed"].append(follow_up["id"])
                    else:
                        supabase.table("lead_follow_ups") \
                            .update({
                                "status": "failed", 
                                "error_message": message
                            }) \
                            .eq("id", follow_up["id"]) \
                            .execute()
                        results["failed"].append(follow_up["id"])
                else:
                    supabase.table("lead_follow_ups") \
                        .update({"status": "failed", "error_message": "Failed to generate content"}) \
                        .eq("id", follow_up["id"]) \
                        .execute()
                    results["failed"].append(follow_up["id"])
                    
            except Exception as e:
                app.logger.error(f"Error processing follow-up {follow_up['id']}: {str(e)}")
                results["failed"].append(follow_up["id"])
        
        return jsonify(results), 200
        
    except Exception as e:
        app.logger.error(f"Error in process_follow_ups: {str(e)}")
        return jsonify({"error": str(e)}), 500



#-------------------------------------------------
#---------the manual dash--------------------
@app.route("/dashboard/manual_email", methods=["GET", "POST"])
def manual_email():
    """Manual email input for users who don't want email forwarding"""
    user_id = _require_user()
    
    if request.method == "POST":
        # Process the manual email
        email_content = request.form.get("email_content", "").strip()
        sender_email = request.form.get("sender_email", "").strip()
        subject = request.form.get("subject", "Inquiry") or "Inquiry"
        
        if not email_content:
            return jsonify({"error": "Email content is required"}), 400
        
        try:
            # Get user's email to use as recipient_email
            user_profile = supabase.table("profiles") \
                .select("email") \
                .eq("id", user_id) \
                .single() \
                .execute().data
            
            user_email = user_profile.get("email") if user_profile else "user@example.com"
            
            # Create email record with ALL required fields
            email_data = {
                "user_id": user_id,
                "sender_email": sender_email or "manual@input.com",
                "recipient_email": user_email,  # REQUIRED FIELD - use user's email
                "original_content": email_content,
                "subject": subject,
                "status": "processing",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "original_user_id": user_id,  # Also set this for consistency
                "is_forwarded": False  # Mark as manual email
            }
            
            # Use SUPABASE_SERVICE (service role) to bypass RLS
            result = SUPABASE_SERVICE.table("emails").insert(email_data).execute()
            
            if result.data:
                # Trigger AI processing
                success = call_edge("/functions/v1/clever-agent/generate-response", 
                                  {"email_ids": [result.data[0]["id"]]})
                
                return jsonify({
                    "success": True,
                    "message": "Email submitted for AI processing",
                    "email_id": result.data[0]["id"]
                })
            else:
                return jsonify({"error": "Failed to save email"}), 500
                
        except Exception as e:
            app.logger.error(f"Error processing manual email: {str(e)}")
            return jsonify({"error": f"Failed to process email: {str(e)}"}), 500
    
    # GET request - render the manual email form
    return render_template("partials/manual_email.html", user_id=user_id)

@app.route("/check_manual_email_status/<email_id>")
def check_manual_email_status(email_id):
    """Check status of manually submitted email"""
    user_id = _require_user()
    
    try:
        # Use service role to bypass RLS for reading
        email = SUPABASE_SERVICE.table("emails") \
            .select("id, status, processed_content, error_message, user_id") \
            .eq("id", email_id) \
            .single() \
            .execute().data
        
        if not email:
            return jsonify({"error": "Email not found"}), 404
            
        # Verify the email belongs to the current user
        if email["user_id"] != user_id:
            return jsonify({"error": "Access denied"}), 403
            
        return jsonify({
            "status": email["status"],
            "processed_content": email.get("processed_content"),
            "error_message": email.get("error_message")
        })
        
    except Exception as e:
        app.logger.error(f"Error checking email status: {str(e)}")
        return jsonify({"error": "Failed to check status"}), 500

#----------------------------------------------------------------
#---------------- manual follow ups -------------------------------

def generate_follow_up_content(lead_id, sequence_step):
    """Generate dynamic, context-aware follow-up content using AI"""
    try:
        app.logger.info(f"Starting dynamic follow-up generation for lead {lead_id}, step {sequence_step}")
        
        # Get lead details with more context
        lead_resp = supabase.table("leads").select("*").eq("id", lead_id).single().execute()
        if not lead_resp.data:
            app.logger.error(f"Lead {lead_id} not found")
            return None
            
        lead = lead_resp.data
        app.logger.info(f"Processing follow-up for: {lead['first_name']} {lead['last_name']} at {lead['email']}")
        
        # Get comprehensive communication history
        previous_emails = supabase.table("emails") \
            .select("subject, original_content, processed_content, sent_at, status") \
            .eq("sender_email", lead["email"]) \
            .or_(f"recipient_email.eq.{lead['email']},sender_email.eq.{lead['email']}") \
            .order("sent_at", desc=True) \
            .limit(10) \
            .execute().data or []
        
        previous_follow_ups = supabase.table("lead_follow_ups") \
            .select("generated_content, sent_at, sequence_step, status") \
            .eq("lead_id", lead_id) \
            .order("sent_at", desc=True) \
            .limit(10) \
            .execute().data or []
        
        app.logger.info(f"Found {len(previous_emails)} emails and {len(previous_follow_ups)} follow-ups")
        
        # Build rich context for AI
        context = f"""
LEAD PROFILE:
- Name: {lead['first_name']} {lead['last_name']}
- Company/Brokerage: {lead.get('brokerage', 'Not specified')}
- Service Interest: {lead.get('service', 'Not specified')}
- Location: {lead.get('city', 'Not specified')}
- Current Status: {lead.get('status', 'new')}
- Initial Contact: {lead.get('email_sent', 'Not recorded')}

FOLLOW-UP CONTEXT:
- This is follow-up #{sequence_step + 1} in the sequence
- Days since initial contact: {FOLLOW_UP_SEQUENCE[sequence_step]['delay_days']}
- Follow-up type: {FOLLOW_UP_SEQUENCE[sequence_step]['name']}

COMMUNICATION HISTORY:
"""

        # Add detailed email history
        if previous_emails:
            context += "\nEMAIL EXCHANGES:\n"
            for i, email in enumerate(previous_emails):
                context += f"\n--- Email {i+1} ({email.get('sent_at', 'Unknown date')}) ---\n"
                context += f"Subject: {email.get('subject', 'No subject')}\n"
                context += f"Status: {email.get('status', 'unknown')}\n"
                
                # Use original content if available, otherwise processed content
                content = email.get('original_content') or email.get('processed_content', 'No content')
                if content and content != 'No content':
                    # Clean and truncate content for context
                    clean_content = ' '.join(content.split()[:100])  # First 100 words
                    context += f"Content: {clean_content}...\n"
        else:
            context += "\nNo previous email exchanges found.\n"

        # Add follow-up history
        if previous_follow_ups:
            context += "\nPREVIOUS FOLLOW-UPS:\n"
            for i, follow_up in enumerate(previous_follow_ups):
                context += f"\n--- Follow-up {i+1} (Step {follow_up['sequence_step']}) ---\n"
                context += f"Sent: {follow_up.get('sent_at', 'Not sent')}\n"
                context += f"Status: {follow_up.get('status', 'unknown')}\n"
                
                content = follow_up.get('generated_content', '')
                if content:
                    clean_content = ' '.join(content.split()[:50])  # First 50 words
                    context += f"Content: {clean_content}...\n"
        else:
            context += "\nNo previous follow-ups sent.\n"

        # Add strategic guidance for the AI
        context += f"""
WRITING INSTRUCTIONS:
- Create a natural, conversational follow-up email
- Reference specific details from the lead's profile and history when relevant
- Adapt tone based on sequence step: earlier steps are more introductory, later steps are more persistent
- Focus on providing value, not just checking in
- Keep it professional but personable
- If this is a later follow-up, acknowledge the previous attempts to connect
- Include a clear call-to-action appropriate for this stage
- Length: 50-150 words, concise but meaningful
- Do NOT use generic templates - make it feel personalized and human

SPECIFIC CONTEXT FOR THIS FOLLOW-UP:
- Sequence position: {sequence_step + 1} of {len(FOLLOW_UP_SEQUENCE)}
- Days since initial contact: {FOLLOW_UP_SEQUENCE[sequence_step]['delay_days']}
- Lead's current engagement level: {'High' if previous_emails else 'Low'}
- Previous interactions: {len(previous_emails)} emails, {len([f for f in previous_follow_ups if f.get('status') == 'sent'])} follow-ups sent

Generate a fresh, non-templated email that builds on this specific context.
"""

        app.logger.info(f"Built comprehensive context for AI (first 500 chars): {context[:500]}...")
        
        # Call AI with enhanced payload
        payload = {
            "context": context,
            "type": "dynamic_follow_up",
            "sequence_step": sequence_step,
            "lead_id": lead_id,
            "lead_name": f"{lead['first_name']} {lead['last_name']}",
            "company": lead.get('brokerage', ''),
            "service_interest": lead.get('service', ''),
            "days_since_contact": FOLLOW_UP_SEQUENCE[sequence_step]['delay_days'],
            "communication_history_count": len(previous_emails),
            "previous_follow_up_count": len([f for f in previous_follow_ups if f.get('status') == 'sent'])
        }
        
        app.logger.info(f"Calling AI with enhanced follow-up payload for step {sequence_step}")
        
        # Use your existing Edge Function call pattern
        response = call_edge("/functions/v1/generate-follow-up", payload, return_response=True)
        
        if response and response.status_code == 200:
            result = response.json()
            content = result.get("content")
            
            if content:
                app.logger.info(f"Successfully generated dynamic follow-up for lead {lead_id}, step {sequence_step}")
                app.logger.debug(f"Generated content: {content[:200]}...")
                return content
            else:
                app.logger.error(f"AI returned empty content for lead {lead_id}")
                return generate_fallback_follow_up(lead, sequence_step)
        else:
            app.logger.error(f"AI call failed for lead {lead_id}: {response.status_code if response else 'No response'}")
            return generate_fallback_follow_up(lead, sequence_step)
            
    except Exception as e:
        app.logger.error(f"Error generating dynamic follow-up content: {str(e)}", exc_info=True)
        return generate_fallback_follow_up(lead, sequence_step) if 'lead' in locals() else None

def generate_fallback_follow_up(lead, sequence_step):
    """Generate a simple fallback follow-up when AI fails"""
    lead_name = lead['first_name'] or "there"
    days = FOLLOW_UP_SEQUENCE[sequence_step]['delay_days']
    
    follow_ups = [
        f"Hi {lead_name}, I wanted to follow up on my previous email about commercial real estate opportunities in your area. Are you still interested in exploring options?",
        f"Hello {lead_name}, checking in to see if you've had a chance to consider commercial properties recently. I'm here to help if you have any questions.",
        f"Hi {lead_name}, I'm following up on our previous conversation about commercial real estate. The market has been active lately - would you like me to update you on current opportunities?",
        f"Hello {lead_name}, I wanted to reconnect regarding commercial property options. Have your requirements changed since we last connected?",
        f"Hi {lead_name}, just checking in to see if you're still in the market for commercial space. I've come across some new listings that might interest you.",
        f"Hello {lead_name}, I'm following up on our previous discussion. Is this still a good time to explore commercial real estate opportunities?"
    ]
    
    # Use sequence step to pick appropriate fallback, or random if beyond list
    return follow_ups[sequence_step % len(follow_ups)]

@app.route("/generate_manual_followups", methods=["POST"])
def generate_manual_followups():
    """Generate AI-powered follow-ups for a manual email"""
    user_id = _require_user()
    
    try:
        data = request.get_json()
        sender_email = data.get("sender_email")
        sender_name = data.get("sender_name")
        subject = data.get("subject")
        email_content = data.get("email_content")
        
        if not all([sender_email, sender_name, email_content]):
            return jsonify({"error": "Missing required fields"}), 400
        
        # Create a temporary lead record for follow-up generation
        lead_data = {
            "user_id": user_id,
            "first_name": sender_name.split()[0] if sender_name else "Lead",
            "last_name": " ".join(sender_name.split()[1:]) if sender_name and " " in sender_name else "Contact",
            "email": sender_email,
            "brokerage": "Unknown",  # Default values
            "service": "Commercial Real Estate",
            "city": "Unknown",
            "status": "new",
            "email_sent": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Insert the lead using service role to bypass RLS
        lead_result = SUPABASE_SERVICE.table("leads").insert(lead_data).execute()
        
        if not lead_result.data:
            return jsonify({"error": "Failed to create lead record"}), 500
            
        lead_id = lead_result.data[0]["id"]
        
        # Also store the original email for context using service role
        email_record = {
            "user_id": user_id,
            "sender_email": sender_email,
            "recipient_email": "manual@input.com",  # Placeholder
            "original_content": email_content,
            "subject": subject,
            "status": "manual_input",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "original_user_id": user_id,
            "is_forwarded": False
        }
        
        SUPABASE_SERVICE.table("emails").insert(email_record).execute()
        
        # Generate follow-ups using the same AI system
        follow_ups = []
        for step, seq in enumerate(FOLLOW_UP_SEQUENCE):
            # Generate content for this follow-up step
            content = generate_follow_up_content(lead_id, step)
            
            if content:
                scheduled_at = datetime.now(timezone.utc) + timedelta(days=seq['delay_days'])
                
                follow_up_data = {
                    "lead_id": lead_id,
                    "sequence_step": step,
                    "scheduled_at": scheduled_at.isoformat(),
                    "status": "pending",
                    "generated_content": content
                }
                
                # Store the follow-up using service role
                follow_up_result = SUPABASE_SERVICE.table("lead_follow_ups").insert(follow_up_data).execute()
                
                if follow_up_result.data:
                    follow_ups.append({
                        "id": follow_up_result.data[0]["id"],
                        "lead_id": lead_id,
                        "day": seq['name'],
                        "date": scheduled_at.strftime("%Y-%m-%d"),
                        "content": content,
                        "sequence_step": step,
                        "delay_days": seq['delay_days']
                    })
        
        return jsonify({
            "success": True,
            "lead_id": lead_id,
            "follow_ups": follow_ups,
            "message": f"Generated {len(follow_ups)} AI-powered follow-ups"
        })
        
    except Exception as e:
        app.logger.error(f"Error generating manual follow-ups: {str(e)}", exc_info=True)
        return jsonify({"error": f"Failed to generate follow-ups: {str(e)}"}), 500






#-----------------------------------------------------------------------------
# Add to app.py
@app.route("/detect_email_settings", methods=["POST"])
def detect_email_settings():
    """Detect email provider settings based on domain"""
    email = request.form.get("email")
    
    if not email or '@' not in email:
        return jsonify({"error": "Invalid email address"}), 400
    
    domain = email.split('@')[1].lower()
    
    # Common email provider settings
    provider_settings = {
        'gmail.com': {
            'smtp_host': 'smtp.gmail.com',
            'smtp_port': 465,
            'imap_host': 'imap.gmail.com',
            'imap_port': 993
        },
        'outlook.com': {
            'smtp_host': 'smtp.office365.com',
            'smtp_port': 587,
            'imap_host': 'outlook.office365.com',
            'imap_port': 993
        },
        'yahoo.com': {
            'smtp_host': 'smtp.mail.yahoo.com',
            'smtp_port': 465,
            'imap_host': 'imap.mail.yahoo.com',
            'imap_port': 993
        },
        'icloud.com': {
            'smtp_host': 'smtp.mail.me.com',
            'smtp_port': 587,
            'imap_host': 'imap.mail.me.com',
            'imap_port': 993
        },
        # Add more providers as needed
    }
    
    # Find matching domain or use generic settings
    settings = None
    for key, value in provider_settings.items():
        if domain == key or domain.endswith(f'.{key}'):
            settings = value
            break
    
    if not settings:
        # Generic settings for unknown domains
        settings = {
            'smtp_host': 'smtp.' + domain,
            'smtp_port': 465,
            'imap_host': 'imap.' + domain,
            'imap_port': 993
        }
    
    return jsonify({
        "email": email,
        "smtp_host": settings['smtp_host'],
        "smtp_port": settings['smtp_port'],
        "imap_host": settings['imap_host'],
        "imap_port": settings['imap_port']
    })
    
from urllib.parse import unquote
             
@app.route('/partials/connect_smtp_form', methods=['GET', 'POST'])
def connect_smtp_form():
    if request.method == 'POST':
        # Handle POST request (if needed)
        pass
    
    # Handle GET request
    user_id = request.args.get('user_id')
    
    # Get and decode the email parameter
    email_param = request.args.get('email', '')
    email = unquote(email_param) if email_param else ''
    
    # Initialize with default values
    smtp_host = "smtp.gmail.com"
    imap_host = "imap.gmail.com"
    smtp_port = 587
    imap_port = 993
    
    # Try to get detected settings from the request
    settings_param = request.args.get('settings')
    print(f"Raw settings parameter: {settings_param}")  # Debug
    
    if settings_param:
        try:
            # URL decode the settings parameter first
            decoded_settings = unquote(settings_param)
            print(f"Decoded settings: {decoded_settings}")  # Debug
            
            settings = json.loads(decoded_settings)
            smtp_host = settings.get('smtp_host', smtp_host)
            imap_host = settings.get('imap_host', imap_host)
            smtp_port = settings.get('smtp_port', smtp_port)
            imap_port = settings.get('imap_port', imap_port)
            print(f"Using detected settings: {settings}")  # For debugging
        except (json.JSONDecodeError, TypeError) as e:
            print(f"Error parsing settings: {e}")  # For debugging
            # If JSON parsing fails, fall back to defaults
            pass
    
    print(f"Final values - Email: {email}, SMTP: {smtp_host}:{smtp_port}, IMAP: {imap_host}:{imap_port}")  # For debugging
    
    return render_template('partials/connect_smtp_form.html', 
                         user_id=user_id, 
                         email=email,
                         smtp_host=smtp_host,
                         imap_host=imap_host,
                         smtp_port=smtp_port,
                         imap_port=imap_port)
    

    
@app.route("/connect_smtp", methods=["POST"])
def connect_smtp():
    user_id = _require_user()
    
    # Check connected accounts limit
    plan = rate_limiter.get_user_plan(user_id)
    account_limit = plan.get('connected_accounts', 1)
    
    if account_limit < 100:  # Not elite plan
        # Count current connected accounts (check if SMTP is configured)
        result = supabase.table("profiles") \
            .select("smtp_enc_password") \
            .eq("id", user_id) \
            .single() \
            .execute()
        
        current_accounts = 1 if result.data and result.data.get('smtp_enc_password') else 0
        
        if current_accounts >= account_limit:
            return jsonify({
                "error": "Account limit reached",
                "message": f"Your {plan['name']} plan allows {account_limit} connected email account(s)",
                "current": current_accounts,
                "limit": account_limit,
                "upgrade_required": account_limit == 1
            }), 403
            
    """Handle SMTP connection setup"""
    try:
        # Get form data
        user_id = request.form.get("user_id")
        email = request.form.get("smtp_email")
        app_password = request.form.get("smtp_password")
        smtp_host = request.form.get("smtp_host")
        smtp_port = request.form.get("smtp_port")
        imap_host = request.form.get("imap_host")
        imap_port = request.form.get("imap_port")
        
        if not all([user_id, email, app_password, smtp_host, smtp_port]):
            return jsonify({"error": "Missing required fields"}), 400
        
        # Encrypt the app password
        encrypted_password = fernet.encrypt(app_password.encode()).decode()
        
        # Update user profile with SMTP settings
        update_data = {
            "smtp_email": email,
            "smtp_enc_password": encrypted_password,
            "smtp_host": smtp_host,
            "smtp_port": int(smtp_port),
            "email": email,  # Also update the main email
            "forwarding_verified": True,  # Mark as manually connected
            "forwarding_verified_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Add IMAP settings if provided
        if imap_host:
            update_data["imap_host"] = imap_host
        if imap_port:
            update_data["imap_port"] = int(imap_port)
        
        # Update the profile
        supabase.table("profiles").update(update_data).eq("id", user_id).execute()
        
        
        app.logger.info(f"SMTP settings saved for user {user_id}, email: {email}")
        
        return jsonify({
            "success": True,
            "message": "SMTP settings saved successfully",
            "redirect_url": f"/dashboard?user_id={user_id}"
        })
        
    except Exception as e:
        app.logger.error(f"Error saving SMTP settings: {str(e)}", exc_info=True)
        return jsonify({"error": f"Failed to save settings: {str(e)}"}), 500
        
        
        
        
        
#---------------------------------------------------------------------------------------
# --- Plan Management Routes ---
@app.route("/api/plan/status", methods=["GET"])
def get_plan_status():
    """Get current plan status and usage"""
    user_id = _require_user()
    plan_info = rate_limiter.get_plan_info(user_id)
    return jsonify(plan_info)

@app.route("/api/plan/upgrade", methods=["POST"])
def upgrade_plan():
    """Handle plan upgrades"""
    user_id = _require_user()
    data = request.get_json()
    new_plan = data.get("plan_name")
    start_trial = data.get("start_trial", False)
    
    if not new_plan or new_plan not in PLANS:
        return jsonify({"error": "Invalid plan name"}), 400
    
    success, message = rate_limiter.update_user_plan(user_id, new_plan, start_trial)
    
    if success:
        return jsonify({
            "success": True,
            "message": message,
            "plan": PLANS[new_plan]
        })
    else:
        return jsonify({"error": message}), 500

@app.route("/api/plan/trial/start", methods=["POST"])
def start_trial():
    """Start free trial for a specific plan"""
    user_id = _require_user()
    data = request.get_json()
    trial_plan = data.get("plan_name", "professional")
    
    if trial_plan not in ['starter', 'professional', 'elite']:
        return jsonify({"error": "Invalid trial plan"}), 400
    
    # Check if user already has an active trial
    try:
        result = supabase.table("profiles") \
            .select("trial_ends_at, subscription_status") \
            .eq("id", user_id) \
            .single() \
            .execute()
        
        if result.data:
            trial_ends_at = result.data.get('trial_ends_at')
            if trial_ends_at:
                trial_ends = datetime.fromisoformat(trial_ends_at.replace('Z', '+00:00'))
                if datetime.now(timezone.utc) < trial_ends:
                    return jsonify({
                        "error": "Trial already active",
                        "message": "You already have an active trial",
                        "trial_ends_at": trial_ends_at
                    }), 400
    except:
        pass
    
    success, message = rate_limiter.update_user_plan(user_id, trial_plan, start_trial=True)
    
    if success:
        plan_info = rate_limiter.get_plan_info(user_id)
        return jsonify({
            "success": True,
            "message": f"Started 14-day free trial of {PLANS[trial_plan]['name']} plan",
            "trial_ends_at": plan_info.get('trial_ends_at'),
            "plan": PLANS[trial_plan]
        })
    else:
        return jsonify({"error": message}), 500
        
        
        
# --- Plan Info Partial for Dashboard ---
@app.route("/dashboard/plan_info")
def dashboard_plan_info():
    """HTMX endpoint to get plan info partial"""
    user_id = _require_user()
    plan_info = rate_limiter.get_plan_info(user_id)
    return render_template("partials/plan_info.html", plan_info=plan_info, user_id=user_id)

# --- Helper to Initialize Trial for New Users ---
def initialize_user_plan(user_id, plan_name="professional", start_trial=True):
    """Initialize plan for new user"""
    try:
        if start_trial:
            success, message = rate_limiter.update_user_plan(user_id, plan_name, start_trial=True)
            if success:
                app.logger.info(f"Initialized {plan_name} trial for user {user_id}")
                return True
        else:
            success, message = rate_limiter.update_user_plan(user_id, plan_name, start_trial=False)
            if success:
                app.logger.info(f"Initialized {plan_name} plan for user {user_id}")
                return True
    except Exception as e:
        app.logger.error(f"Error initializing user plan: {str(e)}")
    
    return False

    
    
def send_email_smtp(from_email, from_password, to_email, subject, body, smtp_host="smtp.gmail.com", smtp_port=465):
    """
    Send an email using SMTP
    """
    try:
        body = clean_placeholders(body)
        # Create message
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject
        
        # Add HTML body
        msg.attach(MIMEText(body, 'html'))
        
        
        # Create secure connection
        context = ssl.create_default_context()
        
        if smtp_port == 465:
            # SSL connection
            server = smtplib.SMTP_SSL(smtp_host, smtp_port, context=context)
            server.login(from_email, from_password)
        else:
            # TLS connection (usually port 587)
            server = smtplib.SMTP(smtp_host, smtp_port)
            server.starttls(context=context)
            server.login(from_email, from_password)
        
        # Send email
        server.send_message(msg)
        server.quit()
        
        app.logger.info(f"Email sent from {from_email} to {to_email}")
        return True
        
    except Exception as e:
        app.logger.error(f"SMTP send failed: {str(e)}")
        raise
    
    
    
import os
import json
import hmac
import hashlib
from flask import Flask, request, jsonify, abort


# You get this from Paddle Dashboard > Developer Tools > Notifications
PADDLE_WEBHOOK_SECRET = os.getenv('PADDLE_WEBHOOK_SECRET', 'pdl_ntfset_...')

def verify_paddle_signature(signature_header, raw_body, secret):
    """Verifies that the webhook came from Paddle."""
    try:
        # 1. Parse the header (format: ts=123;h1=abc)
        parts = dict(item.split('=') for item in signature_header.split(';'))
        timestamp = parts.get('ts')
        received_hash = parts.get('h1')
        
        # 2. Reconstruct the signed payload
        signed_payload = f"{timestamp}:{raw_body.decode('utf-8')}"
        
        # 3. Hash it using your secret
        expected_hash = hmac.new(
            secret.encode('utf-8'),
            signed_payload.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        
        # 4. Compare hashes
        return hmac.compare_digest(expected_hash, received_hash)
    except Exception:
        return False

@app.route('/webhook', methods=['POST'])
def paddle_webhook():
    signature = request.headers.get('Paddle-Signature')
    if not signature:
        abort(401)

    # Use the RAW body for verification
    raw_payload = request.get_data()
    
    if not verify_paddle_signature(signature, raw_payload, PADDLE_WEBHOOK_SECRET):
        print("Invalid signature detected!")
        abort(403)

    data = json.loads(raw_payload)
    event_type = data.get('event_type')
    
    print(f"Webhook received: {event_type}")
    
    # Handle successful payment - listen for both events
    if event_type in ['transaction.completed', 'transaction.updated']:
        try:
            # Get data from the event
            event_data = data.get('data', {})
            
            # Extract custom data
            custom_data = event_data.get('custom_data') or {}
            
            # Get user_id and plan info
            user_id = custom_data.get('userId')
            plan_name = custom_data.get('plan', 'starter')
            billing = custom_data.get('billing', 'monthly')
            
            # Check transaction status
            status = event_data.get('status')
            
            print(f"Extracted data - User ID: {user_id}, Plan: {plan_name}, Billing: {billing}, Status: {status}")
            
            if user_id and status in ['completed', 'paid']:
                print(f"✅ Processing payment for user {user_id}...")
                
                # Update user's plan in database
                success, message = rate_limiter.update_user_plan(user_id, plan_name, start_trial=False)
                
                if success:
                    # Update profile with subscription info (without billing_cycle column)
                    update_data = {
                        "subscription_status": "active"
                    }
                    
                    # Try to update the profile
                    try:
                        supabase.table("profiles").update(update_data).eq("id", user_id).execute()
                        print(f"✅ User {user_id} plan updated to {plan_name}")
                        
                        # Try to send redirect URL via email or store it for the user
                        try:
                            # Store the user_id in a temporary table for login
                            supabase.table("payment_redirects").insert({
                                "user_id": user_id,
                                "created_at": datetime.now(timezone.utc).isoformat(),
                                "status": "pending_login"
                            }).execute()
                            
                            # Send email with login link
                            user_profile = supabase.table("profiles").select("email, full_name").eq("id", user_id).single().execute().data
                            if user_profile and user_profile.get('email'):
                                send_payment_confirmation_email(
                                    user_profile['email'],
                                    user_profile.get('full_name', 'User'),
                                    plan_name,
                                    billing,
                                    user_id
                                )
                        except Exception as e:
                            print(f"Note: Could not send email, but user plan was updated: {e}")
                            
                    except Exception as e:
                        print(f"❌ Database update error: {str(e)}")
                        # Even if database update fails, log the success
                        print(f"💰 Payment processed for user {user_id}, but database update had issues")
                        
                else:
                    print(f"❌ Failed to update user plan: {message}")
                    
                    # Try to notify admin
                    try:
                        admin_email = os.environ.get("ADMIN_EMAIL")
                        if admin_email:
                            send_alert_email(
                                admin_email,
                                f"Payment Processing Error for user {user_id}",
                                f"Plan: {plan_name}, Error: {message}"
                            )
                    except:
                        pass
            else:
                print(f"⚠️ Skipping - Missing user_id or wrong status. User ID: {user_id}, Status: {status}")
                
        except Exception as e:
            print(f"❌ Error processing webhook: {str(e)}")
            app.logger.error(f"Webhook processing error: {str(e)}", exc_info=True)
    
    return jsonify({"status": "success"}), 200
    
@app.route('/payment/success')
def payment_success():
    """Handle successful payment redirect"""
    user_id = request.args.get('user_id')
    transaction_id = request.args.get('transaction_id')
    
    if not user_id:
        return redirect('/app/signin2')
    
    return render_template('/app/dashboard?user_id=${user_id}', 
                         user_id=user_id, 
                         transaction_id=transaction_id)
    
    
import uuid  # Add this import
    
    
@app.route("/auto-register")
def auto_register():
    # 1. Get the parameters from the URL
    user_id = request.args.get("user_id") # This is your lead_id/provided ID
    email = request.args.get("email")
    full_name = request.args.get("full_name", "New User")
    password = "TemporaryPassword123!" # You can also generate a random one

    if not user_id or not email:
        return jsonify({"error": "Missing required parameters: user_id and email"}), 400

    try:
        # 2. CREATE THE AUTH USER (This is what actually authenticates them)
        # We use the Service Role key to bypass email confirmation for auto-registration
        auth_response = SUPABASE_SERVICE.auth.admin.create_user({
            "email": email,
            "password": password,
            "user_metadata": {"full_name": full_name},
            "email_confirm": True # Auto-confirm so they don't have to check email
        })

        # Get the real Supabase Auth UUID
        supabase_uuid = auth_response.user.id

        # 3. CREATE THE PROFILE RECORD
        # We link your lead_id (user_id from params) to the new Auth UUID
        starter_plan = PLANS['starter']
        
        new_profile = {
            "id": supabase_uuid, # Use the actual Auth UUID as the primary key
            "email": email,
            "full_name": full_name,
            "display_name": full_name,
            "plan_name": "starter",
            "subscription_status": "active",
            "monthly_leads_limit": starter_plan['monthly_leads'],
            "monthly_emails_limit": starter_plan['monthly_emails'],
            "monthly_cold_emails_limit": starter_plan['cold_emails'],
            "connected_accounts_limit": starter_plan['connected_accounts'],
            "document_generation_enabled": starter_plan['document_generation'],
            "current_month_leads": 0,
            "current_month_emails": 0,
            "current_month_cold_emails": 0,
            "usage_reset_date": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        supabase.table("profiles").upsert(new_profile).execute()

        # 4. SIGN THEM IN TO GET A SESSION
        # This generates the session so the browser recognizes them as logged in
        login_session = supabase.auth.sign_in_with_password({
            "email": email, 
            "password": password
        })

        # 5. REDIRECT TO YOUR DOMAIN
        # We pass the user_id (the lead id you received) in the URL as requested
        target_url = f"https://replyzeai.com/app/dashboard?user_id={supabase_uuid}"
        return redirect(target_url)

    except Exception as e:
        app.logger.error(f"Authentication Error: {str(e)}")
        # If user already exists, just redirect them
        if "already exists" in str(e).lower():
            return redirect(f"https://replyzeai.com/app/dashboard?user_id={supabase_uuid}")
            
        return jsonify({"error": "Failed to authenticate", "details": str(e)}), 500
        
        
        
        
@app.route("/api/connection/status")
def connection_status():
    """Check email connection status"""
    user_id = _require_user()
    
    try:
        # Check SMTP
        profile = supabase.table("profiles") \
            .select("smtp_enc_password, email") \
            .eq("id", user_id) \
            .single() \
            .execute().data or {}
        
        has_smtp = bool(profile.get("smtp_enc_password"))
        
        # Check Gmail
        gmail_tokens = supabase.table("gmail_tokens") \
            .select("credentials") \
            .eq("user_id", user_id) \
            .execute().data or []
        
        has_gmail = len(gmail_tokens) > 0
        
        return jsonify({
            "user_id": user_id,
            "has_smtp_connection": has_smtp,
            "has_gmail_connection": has_gmail,
            "any_connection": has_smtp or has_gmail,
            "primary_email": profile.get("email"),
            "needs_connection": not (has_smtp or has_gmail)
        })
        
    except Exception as e:
        app.logger.error(f"Error checking connection status: {str(e)}")
        return jsonify({"error": str(e)}), 500
        
        
# ADD this new route to your app.py for marking properties as sold

@app.route("/app/dashboard/properties/<property_id>/mark-sold", methods=["POST"])
def property_mark_sold(property_id):
    """Mark property as sold/closed"""
    user_id = request.form.get("user_id")
    
    try:
        supabase.table("properties") \
            .update({
                "availability_status": "sold",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }) \
            .eq("id", property_id) \
            .eq("user_id", user_id) \
            .execute()
        
        return """
        <script>
            window.location.reload();
        </script>
        """
    except Exception as e:
        print(f"Error marking property as sold: {e}")
        return jsonify({"error": str(e)}), 500
        
    
# ── Final entry point ──
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))